# ==============================================================================
# scripts/Analise/consolidacao_contabil_logic.py
# Módulo de Regras de Negócio: Consolidação e Evolução Contábil
# ==============================================================================

import pandas as pd
import os
import traceback
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.formula.translate import Translator
from openpyxl.styles import Font, PatternFill
import copy
import re

ABAS_SISTEMA = ['Cadastro_Empresas']

def higienizar_nome_aba(nome):
    nome_limpo = re.sub(r'[\\/\?\*\[\]]', '', str(nome))
    return nome_limpo[:31].strip()

def padronizar_conta(conta):
    """ Remove pontos e espaços para garantir comparação correta em formato string contínuo """
    return str(conta).replace('.', '').replace(' ', '').strip()

def copiar_estilo(celula_origem, celula_destino):
    if celula_origem.font: celula_destino.font = copy.copy(celula_origem.font)
    if celula_origem.border: celula_destino.border = copy.copy(celula_origem.border)
    if celula_origem.fill: celula_destino.fill = copy.copy(celula_origem.fill)
    if celula_origem.number_format: celula_destino.number_format = copy.copy(celula_origem.number_format)
    if celula_origem.protection: celula_destino.protection = copy.copy(celula_origem.protection)
    if celula_origem.alignment: celula_destino.alignment = copy.copy(celula_origem.alignment)

def is_formula_estritamente_vertical(formula, coluna_origem='C'):
    if not formula or not isinstance(formula, str): return True
    referencias = re.findall(r'\$?([A-Za-z]+)\$?[0-9]+', formula)
    colunas_encontradas = set(col.upper() for col in referencias)
    for col in colunas_encontradas:
        if col != coluna_origem.upper(): return False
    return True

def exportar_templates(pasta_destino):
    try:
        caminho_params = os.path.join(pasta_destino, "Template_Mestre_Consolidacao.xlsx")
        df_cad = pd.DataFrame([
            {"Empresa": "Empresa Alfa Ltda", "CNPJ": "11.111.111/0001-11", "Arquivo_Balancete": "Balancete_Exemplo_EmpresaAlfa.xlsx", "Plano_Utilizado": "Plano_1"}
        ])
        df_p1 = pd.DataFrame([
            {"Conta_Analitica": "1160101", "Descricao_Conta": "ICMS a Recuperar", "Codigo_Sintetico_BP": "11601"}
        ])

        with pd.ExcelWriter(caminho_params, engine='openpyxl') as writer:
            df_cad.to_excel(writer, sheet_name="Cadastro_Empresas", index=False)
            df_p1.to_excel(writer, sheet_name="Plano_1", index=False)

        wb = openpyxl.load_workbook(caminho_params)
        ws_balanco = wb.create_sheet("Balanço Patrimonial")
        ws_balanco.append(["Código", "Nome da Conta", "Mês Base (Origem de Formatação)"])
        ws_balanco.append(["", "Realizável a Longo Prazo", "=SUM(C3:C7)"])
        for col_idx in range(1, 4):
            ws_balanco.cell(row=1, column=col_idx).font = Font(bold=True)
            ws_balanco.cell(row=2, column=col_idx).font = Font(bold=True)
        wb.save(caminho_params)

        return True, "Templates gerados com sucesso."
    except Exception as e:
        return False, f"Ocorreu um erro: {str(e)}"

def executar_validacao_contas(pasta_balancetes, arquivo_parametros):
    try:
        xls_params = pd.ExcelFile(arquivo_parametros)
        if 'Cadastro_Empresas' not in xls_params.sheet_names:
            return False, "Estrutura inválida: Aba 'Cadastro_Empresas' não localizada."
        return True, "Arquivo de parâmetros estruturalmente válido."
    except Exception as e:
        return False, f"Falha na validação do arquivo: {str(e)}"

def verificar_contas_novas(pasta_balancetes, arquivo_parametros):
    """ Varre os balancetes identificando contas inexistentes na parametrização """
    try:
        xls_params = pd.ExcelFile(arquivo_parametros)
        df_cad = pd.read_excel(xls_params, sheet_name='Cadastro_Empresas')
        
        mapeamento_planos = {}
        for plano in df_cad['Plano_Utilizado'].dropna().unique():
            if plano in xls_params.sheet_names:
                df_plano = pd.read_excel(xls_params, sheet_name=plano, dtype={'Conta_Analitica': str})
                contas_existentes = set(df_plano['Conta_Analitica'].apply(padronizar_conta))
                mapeamento_planos[plano] = contas_existentes

        contas_orfãs = {}

        for idx, row in df_cad.iterrows():
            arquivo_balancete = row.get('Arquivo_Balancete')
            plano_nome = row.get('Plano_Utilizado')
            if pd.isna(arquivo_balancete) or pd.isna(plano_nome): continue
            
            caminho_bal = os.path.join(pasta_balancetes, str(arquivo_balancete))
            if not os.path.exists(caminho_bal): continue
            
            xls_bal = pd.ExcelFile(caminho_bal)
            abas_meses = [aba for aba in xls_bal.sheet_names if str(aba).strip().isdigit() and len(str(aba).strip()) == 2]
            contas_plano_atual = mapeamento_planos.get(plano_nome, set())
            
            for mes in abas_meses:
                df_mes = pd.read_excel(xls_bal, sheet_name=mes, dtype=str)
                df_mes.columns = [str(c).strip().upper() for c in df_mes.columns]
                
                # Identifica dinamicamente a coluna de descrição
                col_descricao = next((c for c in df_mes.columns if c in ['NOME', 'DESCRIÇÃO', 'DESCRICAO', 'DESCRICAO_CONTA']), None)
                col_conta = 'CONTA' if 'CONTA' in df_mes.columns else None
                
                if not col_conta: continue
                
                for _, row_dados in df_mes.iterrows():
                    conta_bruta = str(row_dados.get(col_conta, '')).strip()
                    if not conta_bruta or conta_bruta.lower() == 'nan': continue
                    
                    conta_limpa = padronizar_conta(conta_bruta)
                    if conta_limpa not in contas_plano_atual:
                        descricao = str(row_dados.get(col_descricao, 'N/A')).strip() if col_descricao else 'N/A'
                        chave = (plano_nome, conta_limpa)
                        if chave not in contas_orfãs:
                            contas_orfãs[chave] = {
                                'conta': conta_limpa,
                                'descricao': descricao,
                                'plano': plano_nome
                            }

        lista_retorno = sorted(list(contas_orfãs.values()), key=lambda x: (x['plano'], x['conta']))
        return True, lista_retorno

    except Exception as e:
        return False, f"Erro na auditoria de contas: {str(e)}"

def adicionar_contas_plano(arquivo_parametros, contas_novas):
    """ Insere novas contas na ordem hierárquica (string contínua) e destaca em vermelho """
    try:
        wb = openpyxl.load_workbook(arquivo_parametros)
        fundo_vermelho = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')

        # Agrupar contas por plano para processamento em lote
        contas_por_plano = {}
        for item in contas_novas:
            plano = item['plano']
            if plano not in contas_por_plano: contas_por_plano[plano] = []
            contas_por_plano[plano].append(item)

        for plano, lista_insercao in contas_por_plano.items():
            if plano not in wb.sheetnames: continue
            ws = wb[plano]
            
            # Ordena as contas que serão inseridas (garante alinhamento numérico contínuo via texto)
            lista_insercao = sorted(lista_insercao, key=lambda x: str(x['conta']))
            
            for nova_conta in lista_insercao:
                conta_str = str(nova_conta['conta'])
                linha_insercao = ws.max_row + 1
                
                # Varredura para encontrar a posição hierárquica correta (pulando cabeçalho)
                for row_idx in range(2, ws.max_row + 1):
                    celula_existente = ws.cell(row=row_idx, column=1).value
                    if celula_existente:
                        conta_existente_str = padronizar_conta(celula_existente)
                        if conta_existente_str > conta_str:
                            linha_insercao = row_idx
                            break
                            
                ws.insert_rows(linha_insercao)
                
                c_conta = ws.cell(row=linha_insercao, column=1, value=conta_str)
                c_desc = ws.cell(row=linha_insercao, column=2, value=nova_conta['descricao'])
                c_vinc = ws.cell(row=linha_insercao, column=3, value=nova_conta['vinculo'])
                
                c_conta.fill = fundo_vermelho
                c_desc.fill = fundo_vermelho
                c_vinc.fill = fundo_vermelho

        wb.save(arquivo_parametros)
        return True, "As contas foram incorporadas aos planos correspondentes com sucesso."
    except Exception as e:
        return False, f"Erro ao atualizar arquivo de parâmetros: {str(e)}"

def gerar_relatorios_consolidacao(pasta_balancetes, arquivo_parametros, pasta_destino, gerar_rel1=True, gerar_rel2=True, gerar_rel3=True):
    try:
        # =====================================================================
        # 1. CARREGAMENTO E EXTRAÇÃO (Motor Pandas)
        # =====================================================================
        xls_params = pd.ExcelFile(arquivo_parametros)
        df_cad = pd.read_excel(xls_params, sheet_name='Cadastro_Empresas')
        lista_empresas = df_cad['Empresa'].dropna().unique().tolist()
        
        mapeamento_geral = {}
        for plano in df_cad['Plano_Utilizado'].dropna().unique():
            if plano in xls_params.sheet_names:
                df_plano = pd.read_excel(xls_params, sheet_name=plano, dtype={'Conta_Analitica': str, 'Codigo_Sintetico_BP': str})
                df_plano['Conta_Limpa'] = df_plano['Conta_Analitica'].apply(padronizar_conta)
                df_plano['Codigo_Sintetico_BP'] = df_plano['Codigo_Sintetico_BP'].astype(str).str.strip()
                mapeamento_geral[plano] = dict(zip(df_plano['Conta_Limpa'], df_plano['Codigo_Sintetico_BP']))

        dados_multidimensionais = {}
        meses_detectados = set()

        for idx, row in df_cad.iterrows():
            empresa = str(row.get('Empresa')).strip()
            arquivo_balancete = row.get('Arquivo_Balancete')
            plano_nome = row.get('Plano_Utilizado')
            if pd.isna(arquivo_balancete) or pd.isna(plano_nome): continue
            
            caminho_bal = os.path.join(pasta_balancetes, str(arquivo_balancete))
            if not os.path.exists(caminho_bal): continue
            
            if empresa not in dados_multidimensionais: dados_multidimensionais[empresa] = {}
                
            xls_bal = pd.ExcelFile(caminho_bal)
            abas_meses = [aba for aba in xls_bal.sheet_names if str(aba).strip().isdigit() and len(str(aba).strip()) == 2]
            map_plano = mapeamento_geral.get(plano_nome, {})
            
            for mes in abas_meses:
                mes_str = str(mes).strip()
                meses_detectados.add(mes_str)
                df_mes = pd.read_excel(xls_bal, sheet_name=mes, dtype={'Conta': str})
                
                # Otimização de tratamento de colunas numéricas
                for col in ['Saldo Anterior', 'Movimento', 'Saldo Acumulado']:
                    if col in df_mes.columns:
                        df_mes[col] = pd.to_numeric(df_mes[col], errors='coerce').fillna(0.0)
                    else:
                        df_mes[col] = 0.0

                df_mes['Conta_Limpa'] = df_mes['Conta'].apply(padronizar_conta)
                df_mes['Codigo_Mestre'] = df_mes['Conta_Limpa'].map(map_plano).fillna(df_mes['Conta_Limpa'])
                df_agrupado = df_mes.groupby('Codigo_Mestre')[['Saldo Anterior', 'Movimento', 'Saldo Acumulado']].sum()
                dados_multidimensionais[empresa][mes_str] = df_agrupado.to_dict('index')

        lista_meses = sorted(list(meses_detectados))
        if not lista_meses: return False, "Nenhum período contábil válido identificado nos arquivos."

        # =====================================================================
        # 2. PERSISTÊNCIA VISUAL (Openpyxl)
        # =====================================================================
        wb_template = openpyxl.load_workbook(arquivo_parametros)
        abas_relatorios = [aba for aba in wb_template.sheetnames if aba not in ABAS_SISTEMA and not aba.startswith("Plano_")]

        def processar_aba(ws_origem, ws_destino, tipo_relatorio, mes_ou_empresa_foco=None):
            for col_letter in ['A', 'B', 'C']:
                if col_letter in ws_origem.column_dimensions:
                    ws_destino.column_dimensions[col_letter].width = ws_origem.column_dimensions[col_letter].width

            iterador_dinamico = lista_empresas if tipo_relatorio == 1 else lista_meses

            for row_idx, row in enumerate(ws_origem.iter_rows(values_only=False), start=1):
                if ws_origem.row_dimensions[row_idx].height is not None:
                    ws_destino.row_dimensions[row_idx].height = ws_origem.row_dimensions[row_idx].height

                cell_a_orig, cell_b_orig = row[0], row[1]
                cell_c_orig = row[2] if len(row) > 2 else None
                codigo_conta = padronizar_conta(cell_a_orig.value) if cell_a_orig.value else ""
                
                copiar_estilo(cell_a_orig, ws_destino.cell(row=row_idx, column=1, value=cell_a_orig.value))
                copiar_estilo(cell_b_orig, ws_destino.cell(row=row_idx, column=2, value=cell_b_orig.value))

                tem_formula = cell_c_orig and cell_c_orig.data_type == 'f'
                col_letra_origem = get_column_letter(cell_c_orig.column) if cell_c_orig else 'C'

                for offset, item_dinamico in enumerate(iterador_dinamico):
                    col_sa, col_mov, col_sac = 3 + (offset * 3), 4 + (offset * 3), 5 + (offset * 3)

                    cell_sa = ws_destino.cell(row=row_idx, column=col_sa)
                    cell_mov = ws_destino.cell(row=row_idx, column=col_mov)
                    cell_sac = ws_destino.cell(row=row_idx, column=col_sac)

                    if cell_c_orig:
                        copiar_estilo(cell_c_orig, cell_sa)
                        copiar_estilo(cell_c_orig, cell_mov)
                        copiar_estilo(cell_c_orig, cell_sac)

                    if row_idx == 1:
                        ws_destino.column_dimensions[get_column_letter(col_sa)].width = 18
                        ws_destino.column_dimensions[get_column_letter(col_mov)].width = 18
                        ws_destino.column_dimensions[get_column_letter(col_sac)].width = 18
                        ws_destino.column_dimensions[get_column_letter(col_sa)].hidden = True
                        ws_destino.column_dimensions[get_column_letter(col_sac)].hidden = True

                        if tipo_relatorio == 1:
                            cell_sa.value, cell_mov.value, cell_sac.value = f"{item_dinamico} - Anterior", f"{item_dinamico} - Movimento", f"{item_dinamico} - Final"
                        else:
                            cell_sa.value, cell_mov.value, cell_sac.value = f"Mês {item_dinamico} - Anterior", f"Mês {item_dinamico} - Movimento", f"Mês {item_dinamico} - Final"
                        continue

                    if not tem_formula and codigo_conta and codigo_conta != "None":
                        sa, mov, sac = 0.0, 0.0, 0.0

                        if tipo_relatorio == 1: 
                            dados_conta = dados_multidimensionais.get(item_dinamico, {}).get(mes_ou_empresa_foco, {}).get(codigo_conta, {})
                            sa, mov, sac = dados_conta.get('Saldo Anterior', 0), dados_conta.get('Movimento', 0), dados_conta.get('Saldo Acumulado', 0)
                        elif tipo_relatorio == 2:
                            for emp in lista_empresas:
                                dados_conta = dados_multidimensionais.get(emp, {}).get(item_dinamico, {}).get(codigo_conta, {})
                                sa += dados_conta.get('Saldo Anterior', 0)
                                mov += dados_conta.get('Movimento', 0)
                                sac += dados_conta.get('Saldo Acumulado', 0)
                        elif tipo_relatorio == 3: 
                            dados_conta = dados_multidimensionais.get(mes_ou_empresa_foco, {}).get(item_dinamico, {}).get(codigo_conta, {})
                            sa, mov, sac = dados_conta.get('Saldo Anterior', 0), dados_conta.get('Movimento', 0), dados_conta.get('Saldo Acumulado', 0)

                        if sa != 0: cell_sa.value = sa
                        if mov != 0: cell_mov.value = mov
                        if sac != 0: cell_sac.value = sac

                    elif tem_formula:
                        formula_original = cell_c_orig.value
                        if is_formula_estritamente_vertical(formula_original, col_letra_origem):
                            cell_sa.value = Translator(formula_original, origin=f"{col_letra_origem}{row_idx}").translate_formula(f"{get_column_letter(col_sa)}{row_idx}")
                            cell_mov.value = Translator(formula_original, origin=f"{col_letra_origem}{row_idx}").translate_formula(f"{get_column_letter(col_mov)}{row_idx}")
                            cell_sac.value = Translator(formula_original, origin=f"{col_letra_origem}{row_idx}").translate_formula(f"{get_column_letter(col_sac)}{row_idx}")
                        else:
                            cell_sa.value = "FÓRMULA INVÁLIDA"
                            cell_mov.value = "FÓRMULA INVÁLIDA"
                            cell_sac.value = "FÓRMULA INVÁLIDA"

        # Relatório 1
        if gerar_rel1:
            wb_rel1 = openpyxl.Workbook(); wb_rel1.remove(wb_rel1.active)
            for mes in lista_meses:
                for nome_aba_tmpl in abas_relatorios:
                    ws_dest = wb_rel1.create_sheet(title=higienizar_nome_aba(f"{mes}_{nome_aba_tmpl}"))
                    processar_aba(wb_template[nome_aba_tmpl], ws_dest, 1, mes)
            wb_rel1.save(os.path.join(pasta_destino, "Relatorio_1_Comparativo_Empresas_Por_Mes.xlsx"))

        # Relatório 2
        if gerar_rel2:
            wb_rel2 = openpyxl.Workbook(); wb_rel2.remove(wb_rel2.active)
            for nome_aba_tmpl in abas_relatorios:
                ws_dest = wb_rel2.create_sheet(title=higienizar_nome_aba(nome_aba_tmpl))
                processar_aba(wb_template[nome_aba_tmpl], ws_dest, 2)
            wb_rel2.save(os.path.join(pasta_destino, "Relatorio_2_Consolidado_Global.xlsx"))

        # Relatório 3
        if gerar_rel3:
            wb_rel3 = openpyxl.Workbook(); wb_rel3.remove(wb_rel3.active)
            for empresa in lista_empresas:
                for nome_aba_tmpl in abas_relatorios:
                    ws_dest = wb_rel3.create_sheet(title=higienizar_nome_aba(f"{empresa}_{nome_aba_tmpl}"))
                    processar_aba(wb_template[nome_aba_tmpl], ws_dest, 3, empresa)
            wb_rel3.save(os.path.join(pasta_destino, "Relatorio_3_Evolucao_Individual_Empresas.xlsx"))

        return True, "Extração e consolidação concluídas com sucesso. Relatórios salvos no diretório de destino."

    except Exception as e:
        return False, f"Falha na execução: {str(e)}\n\nDetalhes:\n{traceback.format_exc()}"