import os
from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QLabel, QPushButton, 
                             QFileDialog, QCheckBox, QGroupBox, QHBoxLayout, 
                             QMessageBox, QProgressBar, QScrollArea, QGridLayout, QLineEdit, QApplication, QDialog, QTextBrowser)
from PyQt6.QtCore import Qt

# Importando a função do Motor Lógico
from scripts.Contabilidade.analise_evolucao_logic import executar_analise_evolucao

APP_STYLESHEET = """
QWidget#MainWindow { background-color: #F1F3F4; }
QWidget#ScrollContent { background-color: #FFFFFF; }
QWidget#Footer { background-color: #FFFFFF; border-top: 1px solid #DADCE0; }
QLabel#AppTitle { font-size: 22px; font-weight: 600; color: #1A73E8; padding: 5px 0; }
QLabel#SectionTitle { font-size: 14px; font-weight: bold; color: #202124; margin-top: 15px; margin-bottom: 5px; }
QLabel#GridHeader { font-size: 12px; font-weight: bold; color: #5F6368; padding-bottom: 5px; }
QLabel { font-size: 12px; color: #3C4043; }

QGroupBox { background-color: #FFFFFF; border: 1px solid #E1E4E8; border-radius: 8px; font-size: 13px; font-weight: 600; color: #202124; margin-top: 15px; padding-top: 15px; }
QGroupBox::title { subcontrol-origin: margin; subcontrol-position: top left; padding: 0 8px; left: 15px; }

/* Botões Principais no Rodapé */
QPushButton#btn_processar { background-color: #1A73E8; color: white; font-size: 13px; font-weight: 600; border: none; padding: 10px 20px; border-radius: 6px; }
QPushButton#btn_processar:hover { background-color: #1557B0; }
QPushButton#btn_reprocessar { background-color: #34A853; color: white; font-size: 13px; font-weight: 600; border: none; padding: 10px 20px; border-radius: 6px; }
QPushButton#btn_reprocessar:hover { background-color: #2b8c44; }
QPushButton:disabled { background-color: #A0C3FF; }

/* Botão Secundário (Procurar Arquivo) */
QPushButton#btn_secundario { background-color: #F8F9FA; color: #3C4043; font-weight: bold; border: 1px solid #DADCE0; padding: 7px 15px; border-radius: 4px; }
QPushButton#btn_secundario:hover { background-color: #E8EAED; }

/* Botão Ajuda (Wiki) */
QPushButton#btn_help { 
    background-color: transparent; border: 2px solid #1A73E8; 
    color: #1A73E8; font-size: 16px; font-weight: bold; 
    border-radius: 15px; min-width: 30px; min-height: 30px; max-width: 30px; max-height: 30px; 
}
QPushButton#btn_help:hover { background-color: #E8F0FE; }

QScrollArea { border: none; background-color: transparent; }

/* Inputs Elásticos */
QLineEdit { background-color: #F8F9FA; border: 1px solid #DADCE0; border-radius: 4px; padding: 6px; color: #202124; font-weight: bold; }
QLineEdit:focus { border: 1px solid #1A73E8; background-color: #FFFFFF; }
QLineEdit[readOnly="true"] { background-color: #F1F3F4; color: #5F6368; font-weight: normal; border: 1px solid #E8EAED; }

/* Barra de Progresso Inteligente */
QProgressBar { border: 1px solid #DADCE0; border-radius: 3px; background-color: #F1F3F4; text-align: center; }
QProgressBar::chunk { background-color: #1A73E8; border-radius: 3px; }
QProgressBar[state="success"]::chunk { background-color: #34A853; }
"""

class AjudaDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Wiki: Análise de Evolução")
        self.resize(700, 550)
        self.setStyleSheet(APP_STYLESHEET)
        
        lay = QVBoxLayout(self)
        lay.setContentsMargins(25, 25, 25, 25)
        lay.setSpacing(15)
        
        lbl_title = QLabel("Documentação da Ferramenta", objectName="AppTitle")
        lay.addWidget(lbl_title)
        
        txt_wiki = QTextBrowser()
        txt_wiki.setHtml("""
        <div style='font-family: Calibri, sans-serif; font-size: 14px; color: #3C4043; line-height: 1.5;'>
            <h3 style='color: #1A73E8;'>Visão Geral</h3>
            <p>Este sistema consolida balancetes mensais de clientes, gera o <b>Balancete Histórico</b> auditável e constrói de forma automática as planilhas <b>PTA (Plano de Trabalho de Auditoria)</b>. A aplicação incorpora inteligência matemática para identificar anomalias, cruzando movimentações, saldos e desvios de média, aderindo às normas da <b>NBC TA 530</b>.</p>
            
            <h3 style='color: #1A73E8;'>Estrutura do Arquivo Requerido</h3>
            <p>O arquivo Excel original deve conter estritamente:</p>
            <ul>
                <li>Uma aba chamada <b>'Parametros'</b> com o Plano de Contas referenciado.</li>
                <li>Abas mensais numeradas de <b>'1' a '12'</b> contendo o balancete bruto do cliente.</li>
            </ul>

            <h3 style='color: #1A73E8;'>Regras de Negócio e Funcionalidades</h3>
            <ul>
                <li><b>Inclusão Inteligente:</b> Identifica saldos em contas não mapeadas no parâmetro e as adiciona automaticamente sugerindo classificação Sintética (S) ou Analítica (A).</li>
                <li><b>Parametrização NBC TA 530:</b> O sistema varre as contas mapeadas. Se a conta for de Ativo/Passivo e atingir um limite percentual estipulado, ela entra na amostra do PTA. Para Resultados, além do saldo acumulado, averigua-se a quebra de padrão e desvio em relação à média mensal dos meses vivos da conta.</li>
                <li><b>Atualizar Balancete Histórico:</b> Após um processamento inicial, você pode alterar classificações na aba de Parâmetros do Excel gerado. Clicar em "Atualizar" reprocessará a matemática e as regras sem quebrar a estrutura.</li>
            </ul>
        </div>
        """)
        txt_wiki.setStyleSheet("border: 1px solid #DADCE0; border-radius: 6px; background-color: #F8F9FA;")
        lay.addWidget(txt_wiki)
        
        lay_botoes = QHBoxLayout()
        btn_fechar = QPushButton("Entendi")
        btn_fechar.setObjectName("btn_secundario")
        btn_fechar.clicked.connect(self.accept)
        
        btn_exportar = QPushButton("Exportar Arquivo Exemplo (Template)")
        btn_exportar.setObjectName("btn_processar")
        btn_exportar.clicked.connect(self.exportar_exemplo)
        
        lay_botoes.addStretch()
        lay_botoes.addWidget(btn_fechar)
        lay_botoes.addWidget(btn_exportar)
        lay.addLayout(lay_botoes)

    def exportar_exemplo(self):
        caminho, _ = QFileDialog.getSaveFileName(self, "Salvar Template", "Template_Balancete_Cliente.xlsx", "Excel Files (*.xlsx)")
        if not caminho: return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            wb = Workbook()
            
            # Aba Parametros
            ws_param = wb.active
            ws_param.title = "Parametros"
            headers_p = ['Chave Cliente', 'Chave D&M', 'Classificação', 'Descrição', 'Sint./An.', 'At/Pas/Res', 'Indice']
            for c, val in enumerate(headers_p, 1):
                ws_param.cell(1, c, val).font = Font(bold=True, color="FFFFFF")
                ws_param.cell(1, c).fill = PatternFill(start_color="1A73E8", fill_type="solid")
            
            dados_p = [
                ['1', '1', '1', 'ATIVO', 'S', 'A', 1],
                ['111', '101', '1.1.1', 'CAIXA E EQUIVALENTES', 'S', 'A', 2],
                ['11101', '10101', '1.1.1.01', 'CAIXA GERAL', 'A', 'A', 3],
                ['3', '3', '3', 'RECEITAS', 'S', 'R', 4],
                ['301', '301', '3.1.1', 'RECEITA DE VENDAS', 'S', 'R', 5],
                ['30101', '30101', '3.1.1.01', 'VENDAS DE MERCADORIAS', 'A', 'R', 6]
            ]
            for r, linha in enumerate(dados_p, 2):
                for c, val in enumerate(linha, 1):
                    ws_param.cell(r, c, val)
                    
            # Aba 1 (Janeiro)
            ws_jan = wb.create_sheet("1")
            headers_m = ['Atividade', 'Conta', 'Descrição', 'Cod. Reduzido', 'Saldo Anterior', 'Débito', 'Crédito', 'Movimento', 'Saldo Acumulado']
            for c, val in enumerate(headers_m, 1):
                ws_jan.cell(1, c, val).font = Font(bold=True)
                
            dados_jan = [
                ['Geral', '11101', 'CAIXA GERAL', '1.1.1.01', 5000.00, 2000.00, 0.00, 2000.00, 7000.00],
                ['Geral', '30101', 'VENDAS DE MERCADORIAS', '3.1.1.01', 0.00, 0.00, 2000.00, -2000.00, -2000.00]
            ]
            for r, linha in enumerate(dados_jan, 2):
                for c, val in enumerate(linha, 1):
                    ws_jan.cell(r, c, val)
                    
            wb.save(caminho)
            QMessageBox.information(self, "Sucesso", "Template gerado com sucesso!\nUse este arquivo como modelo para estruturar os dados reais.")
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Ocorreu um erro ao gerar o modelo:\n{e}")

class AnaliseEvolucaoApp(QWidget):
    def __init__(self):
        super().__init__()
        self.caminho_arquivo = ""
        self.ultimo_pta_gerado = ""
        self.initUI()

    def initUI(self):
        self.setObjectName("MainWindow")
        self.setStyleSheet(APP_STYLESHEET)
        
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        
        # -------------------------------------------------------------
        # ÁREA SUPERIOR (Rolável)
        # -------------------------------------------------------------
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        content = QWidget()
        content.setObjectName("ScrollContent")
        layout = QVBoxLayout(content)
        layout.setContentsMargins(30, 20, 30, 30)
        layout.setSpacing(20)
        
        # Header com Botão Wiki (?)
        lay_header = QHBoxLayout()
        lbl_title = QLabel("Análise de Evolução", objectName="AppTitle")
        btn_ajuda = QPushButton("?")
        btn_ajuda.setObjectName("btn_help")
        btn_ajuda.setToolTip("Abrir documentação e wiki da ferramenta")
        btn_ajuda.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_ajuda.clicked.connect(self.abrir_wiki)
        
        lay_header.addWidget(lbl_title)
        lay_header.addStretch()
        lay_header.addWidget(btn_ajuda)
        layout.addLayout(lay_header)

        # GRUPO 1: Arquivos
        g_arq = QGroupBox("1. Seleção de Arquivo")
        lay_arq = QVBoxLayout()
        lay_arq.setContentsMargins(20, 20, 20, 20)
        lay_arq.setSpacing(15)
        
        lay_file = QHBoxLayout()
        self.txt_arquivo = QLineEdit()
        self.txt_arquivo.setReadOnly(True)
        self.txt_arquivo.setPlaceholderText("Selecione o arquivo com os balancetes mensais...")
        
        btn_arq = QPushButton("Procurar...")
        btn_arq.setObjectName("btn_secundario")
        btn_arq.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_arq.clicked.connect(self.selecionar_arquivo)
        
        lay_file.addWidget(self.txt_arquivo, stretch=1)
        lay_file.addWidget(btn_arq)
        
        self.chk_inclusao = QCheckBox("Habilitar Inclusão e Classificação Automática de Contas Órfãs")
        self.chk_inclusao.setChecked(True)
        self.chk_inclusao.setStyleSheet("font-weight: normal; color: #3C4043;")
        
        lay_arq.addLayout(lay_file)
        lay_arq.addWidget(self.chk_inclusao)
        g_arq.setLayout(lay_arq)
        layout.addWidget(g_arq)

        # GRUPO 2: Escopo e Parâmetros
        g_escopo = QGroupBox("2. Parametrização de Risco (Limites de Auditoria)")
        lay_escopo = QVBoxLayout()
        lay_escopo.setContentsMargins(20, 20, 20, 20)
        lay_escopo.setSpacing(15)
        
        lay_check = QHBoxLayout()
        lay_check.setSpacing(30)
        self.chk_ativo = QCheckBox("PTA - Ativo"); self.chk_ativo.setChecked(True)
        self.chk_passivo = QCheckBox("PTA - Passivo"); self.chk_passivo.setChecked(True)
        self.chk_resultado = QCheckBox("PTA - Resultado"); self.chk_resultado.setChecked(True)
        lay_check.addWidget(self.chk_ativo)
        lay_check.addWidget(self.chk_passivo)
        lay_check.addWidget(self.chk_resultado)
        lay_check.addStretch()
        lay_escopo.addLayout(lay_check)

        self.container_ativo, self.regras_ativo_ui = self.criar_formulario_basico("REGRAS ATIVO:", "Chave D&M")
        self.container_passivo, self.regras_passivo_ui = self.criar_formulario_basico("REGRAS PASSIVO:", "Chave Cliente")
        self.container_resultado, self.regras_resultado_ui = self.criar_formulario_composto("REGRAS RESULTADO:")
        
        lay_escopo.addWidget(self.container_ativo)
        lay_escopo.addWidget(self.container_passivo)
        lay_escopo.addWidget(self.container_resultado)
        
        self.chk_ativo.toggled.connect(self.container_ativo.setVisible)
        self.chk_passivo.toggled.connect(self.container_passivo.setVisible)
        self.chk_resultado.toggled.connect(self.container_resultado.setVisible)
        
        g_escopo.setLayout(lay_escopo)
        layout.addWidget(g_escopo)
        layout.addStretch()
        
        scroll.setWidget(content)
        main_layout.addWidget(scroll, stretch=1)

        # -------------------------------------------------------------
        # RODAPÉ FIXO (Sempre Visível)
        # -------------------------------------------------------------
        footer = QWidget()
        footer.setObjectName("Footer")
        lay_footer = QHBoxLayout(footer)
        lay_footer.setContentsMargins(30, 15, 30, 15)
        lay_footer.setSpacing(20)

        lay_status = QVBoxLayout()
        lay_status.setSpacing(5)
        self.lbl_status = QLabel("Aguardando processamento...")
        self.lbl_status.setStyleSheet("color: #5F6368; font-size: 11px;")
        
        self.barra_progresso = QProgressBar()
        self.barra_progresso.setFixedHeight(6)
        self.barra_progresso.setTextVisible(False)
        self.barra_progresso.setFixedWidth(250)
        self.barra_progresso.setProperty("state", "normal") # Estado dinâmico do CSS
        
        lay_status.addWidget(self.lbl_status)
        lay_status.addWidget(self.barra_progresso)
        
        self.btn_processar = QPushButton("Processar Base Original")
        self.btn_processar.setObjectName("btn_processar")
        self.btn_processar.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_processar.clicked.connect(lambda: self.processar_dados(is_reprocess=False))
        
        self.btn_reprocessar = QPushButton("Atualizar Balancete Histórico")
        self.btn_reprocessar.setObjectName("btn_reprocessar")
        self.btn_reprocessar.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_reprocessar.clicked.connect(lambda: self.processar_dados(is_reprocess=True))

        lay_footer.addLayout(lay_status)
        lay_footer.addStretch()
        lay_footer.addWidget(self.btn_processar)
        lay_footer.addWidget(self.btn_reprocessar)

        main_layout.addWidget(footer)

        self.resize(1000, 750)
        self.setWindowTitle("Análise de Evolução - Auditoria")

    # ==============================================================
    # WIKI E MÉTODOS DE TELA
    # ==============================================================
    def abrir_wiki(self):
        dlg = AjudaDialog(self)
        dlg.exec()

    def criar_formulario_basico(self, titulo, nome_chave):
        container = QWidget()
        lay = QVBoxLayout(container)
        lay.setContentsMargins(0, 10, 0, 10)
        
        lbl = QLabel(titulo)
        lbl.setObjectName("SectionTitle")
        lay.addWidget(lbl)
        
        grid = QGridLayout()
        grid.setSpacing(10)
        
        headers = ["Faixa", "Saldo Mín. >", "Saldo Máx. Até", nome_chave]
        for col, text in enumerate(headers):
            lbl_h = QLabel(text)
            lbl_h.setObjectName("GridHeader")
            lbl_h.setAlignment(Qt.AlignmentFlag.AlignCenter)
            grid.addWidget(lbl_h, 0, col)
            
        mins = ["0,10%", "0,30%", "0,60%", "0,90%", "1,20%", "1,50%", "1,80%", "2,00%"]
        maxs = ["0,30%", "0,60%", "0,90%", "1,20%", "1,50%", "1,80%", "2,00%", ""]
        lista_inputs = []
        
        for i in range(8):
            lbl_faixa = QLabel(f"{i+1}")
            lbl_faixa.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lbl_faixa.setStyleSheet("font-weight: bold; color: #5F6368;")
            
            inp_min = QLineEdit(mins[i]); inp_min.setAlignment(Qt.AlignmentFlag.AlignCenter)
            inp_max = QLineEdit(maxs[i]); inp_max.setAlignment(Qt.AlignmentFlag.AlignCenter)
            inp_chave = QLineEdit(""); inp_chave.setAlignment(Qt.AlignmentFlag.AlignCenter)
            
            grid.addWidget(lbl_faixa, i+1, 0)
            grid.addWidget(inp_min, i+1, 1)
            grid.addWidget(inp_max, i+1, 2)
            grid.addWidget(inp_chave, i+1, 3)
            
            lista_inputs.append({'min': inp_min, 'max': inp_max, 'chave': inp_chave})
            
        grid.setColumnStretch(1, 1)
        grid.setColumnStretch(2, 1)
        grid.setColumnStretch(3, 3)
        
        lay.addLayout(grid)
        return container, lista_inputs

    def criar_formulario_composto(self, titulo):
        container = QWidget()
        lay = QVBoxLayout(container)
        lay.setContentsMargins(0, 10, 0, 10)
        
        lbl = QLabel(titulo)
        lbl.setObjectName("SectionTitle")
        lay.addWidget(lbl)
        
        grid = QGridLayout()
        grid.setSpacing(10)
        
        headers = ["Faixa", "Saldo Mín. >", "Saldo Máx. Até", "Chave D&M", "Movimento(mês) >= % Média"]
        for col, text in enumerate(headers):
            lbl_h = QLabel(text)
            lbl_h.setObjectName("GridHeader")
            lbl_h.setAlignment(Qt.AlignmentFlag.AlignCenter)
            grid.addWidget(lbl_h, 0, col)
            
        mins = ["0,20%", "0,40%", "0,70%", "1,00%", "1,30%", "1,60%", "1,90%", "2,10%"]
        maxs = ["0,40%", "0,70%", "1,00%", "1,30%", "1,60%", "1,90%", "2,10%", ""]
        percs = ["100%", "75%", "50%", "25%", "20%", "15%", "10%", "5%"]
        lista_inputs = []
        
        for i in range(8):
            lbl_faixa = QLabel(f"{i+1}")
            lbl_faixa.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lbl_faixa.setStyleSheet("font-weight: bold; color: #5F6368;")
            
            inp_min = QLineEdit(mins[i]); inp_min.setAlignment(Qt.AlignmentFlag.AlignCenter)
            inp_max = QLineEdit(maxs[i]); inp_max.setAlignment(Qt.AlignmentFlag.AlignCenter)
            inp_chave = QLineEdit("30101"); inp_chave.setAlignment(Qt.AlignmentFlag.AlignCenter)
            inp_perc = QLineEdit(percs[i]); inp_perc.setAlignment(Qt.AlignmentFlag.AlignCenter)
            
            grid.addWidget(lbl_faixa, i+1, 0)
            grid.addWidget(inp_min, i+1, 1)
            grid.addWidget(inp_max, i+1, 2)
            grid.addWidget(inp_chave, i+1, 3)
            grid.addWidget(inp_perc, i+1, 4)
            
            lista_inputs.append({'min': inp_min, 'max': inp_max, 'chave': inp_chave, 'perc': inp_perc})
            
        grid.setColumnStretch(1, 1)
        grid.setColumnStretch(2, 1)
        grid.setColumnStretch(3, 2)
        grid.setColumnStretch(4, 1)
        
        lay.addLayout(grid)
        return container, lista_inputs

    def selecionar_arquivo(self):
        caminho, _ = QFileDialog.getOpenFileName(self, "Selecionar Arquivo Original", "", "Excel Files (*.xlsx *.xlsb)")
        if caminho:
            self.caminho_arquivo = caminho
            self.txt_arquivo.setText(caminho)
            self.lbl_status.setText("Pronto para processar.")

    def set_estado_barra(self, estado):
        """ Atualiza dinamicamente a cor da barra de progresso alterando a classe via CSS """
        self.barra_progresso.setProperty("state", estado)
        self.barra_progresso.style().unpolish(self.barra_progresso)
        self.barra_progresso.style().polish(self.barra_progresso)

    def processar_dados(self, is_reprocess=False):
        if not self.caminho_arquivo:
            QMessageBox.warning(self, "Aviso", "Por favor, selecione o arquivo Excel Original primeiro (Base de saldos).")
            return

        opcoes = {
            'ativo': self.chk_ativo.isChecked(),
            'passivo': self.chk_passivo.isChecked(),
            'resultado': self.chk_resultado.isChecked(),
            'inclusao_inteligente': self.chk_inclusao.isChecked(),
            'is_reprocess': is_reprocess
        }

        if is_reprocess:
            caminho_pta_alvo = ""
            if self.ultimo_pta_gerado and os.path.exists(self.ultimo_pta_gerado):
                resp = QMessageBox.question(self, "Atualizar", 
                    f"Deseja reprocessar usando os parâmetros do último arquivo gerado?\n\n{os.path.basename(self.ultimo_pta_gerado)}\n\n(Clique em 'No' para buscar outro arquivo).",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
                if resp == QMessageBox.StandardButton.Yes:
                    caminho_pta_alvo = self.ultimo_pta_gerado
            
            if not caminho_pta_alvo:
                caminho_pta_alvo, _ = QFileDialog.getOpenFileName(self, "Selecionar Arquivo PTA Editado", "", "Excel Files (*.xlsx *.xlsb)")
                if not caminho_pta_alvo: return
                
            opcoes['caminho_pta_reprocess'] = caminho_pta_alvo

        nome_sugerido = self.caminho_arquivo.replace(".xlsx", "_Analise_Evolucao_PTA.xlsx").replace(".xlsb", "_Analise_Evolucao_PTA.xlsx")
        caminho_saida, _ = QFileDialog.getSaveFileName(self, "Salvar Relatório Como...", nome_sugerido, "Excel Files (*.xlsx)")
        if not caminho_saida: return
        
        opcoes['caminho_saida'] = caminho_saida

        regras = {
            'ativo': [{'min': r['min'].text(), 'max': r['max'].text(), 'chave': r['chave'].text()} for r in self.regras_ativo_ui],
            'passivo': [{'min': r['min'].text(), 'max': r['max'].text(), 'chave': r['chave'].text()} for r in self.regras_passivo_ui],
            'resultado': [{'min': r['min'].text(), 'max': r['max'].text(), 'chave': r['chave'].text(), 'perc': r['perc'].text()} for r in self.regras_resultado_ui]
        }

        try:
            self.lbl_status.setText("Lendo bases e calculando amostragem...")
            self.btn_processar.setEnabled(False)
            self.btn_reprocessar.setEnabled(False)
            self.set_estado_barra("normal")
            
            # Atualização forçada da tela para o usuário ver que começou
            self.barra_progresso.setValue(45) 
            QApplication.processEvents() 
            
            arquivo_gerado, log_erros = executar_analise_evolucao(self.caminho_arquivo, opcoes, regras)
            self.ultimo_pta_gerado = arquivo_gerado
            
            # Pinta a barra de verde (Estado Success)
            self.set_estado_barra("success")
            self.barra_progresso.setValue(100)
            self.lbl_status.setText("Processo finalizado com sucesso!")
            
            msg = f"Auditoria e Balancete Histórico gerados com sucesso!\n\nSalvo em:\n{os.path.basename(arquivo_gerado)}"
            
            # O sistema trava aqui até o usuário clicar em OK na caixinha de Sucesso
            QMessageBox.information(self, "Sucesso", msg) 
            
        except Exception as e:
            self.lbl_status.setText("Processo abortado devido a erro crítico.")
            QMessageBox.critical(self, "Erro Crítico", f"Ocorreu um erro durante o processamento:\n{str(e)}")
            
        finally:
            # Após o usuário clicar no OK, a barra esvazia e volta a ficar azul para a próxima tarefa
            self.set_estado_barra("normal")
            self.barra_progresso.setValue(0)
            self.lbl_status.setText("Aguardando novo processamento...")
            self.btn_processar.setEnabled(True)
            self.btn_reprocessar.setEnabled(True)

def main():
    janela = AnaliseEvolucaoApp()
    janela.show()
    return None, janela