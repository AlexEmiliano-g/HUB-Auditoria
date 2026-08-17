import os

import pandas as pd
import re

def obter_nome_aba_seguro(caminho_arquivo, abas_existentes):
    """
    Gera o nome da aba para o Excel seguindo as regras do HUB:
    1. Se começar com 'B_XX', tenta usar apenas os dígitos (ex: '01').
    2. Se a aba 'XX' já existir (colisão), utiliza o nome completo do arquivo.
    3. Se não tiver o padrão 'B_XX', utiliza o nome completo.
    4. Limita a 31 caracteres (limite nativo da biblioteca openpyxl e do Excel).
    5. Segurança extra: se até o nome completo gerar colisão, adiciona um sufixo numérico.
    """
    import os
    import re
    
    nome_base = os.path.basename(caminho_arquivo)
    nome_sem_extensao = os.path.splitext(nome_base)[0]
    
    # 1. Tenta extrair o padrão B_XX
    match = re.match(r"B_(\d{2})", nome_base, re.IGNORECASE)
    
    if match:
        tentativa_nome = match.group(1)
        # 2. Se já existir, aplica a regra de usar o nome completo sem extensão
        if tentativa_nome in abas_existentes:
            tentativa_nome = nome_sem_extensao
    else:
        # 3. Não tem padrão B_, usa o nome original
        tentativa_nome = nome_sem_extensao
        
    # 4. Limita aos 31 caracteres do Excel
    nome_final = tentativa_nome[:31]
    
    # 5. Fallback final de segurança absoluta 
    # (Ex: caso o usuário suba dois arquivos idênticos B_01Teste de pastas diferentes)
    contador = 1
    nome_original = nome_final
    while nome_final in abas_existentes:
        sufixo = f"_{contador}"
        nome_final = f"{nome_original[:31-len(sufixo)]}{sufixo}"
        contador += 1
        
    return nome_final

# ==============================================================================
# FUNÇÕES AUXILIARES
# ==============================================================================

def _converter_valor_monetario(valor):
    """
    Converte um valor monetário para float.

    Reconhece indicadores contábeis:
        D = valor positivo
        C = valor negativo

    Exemplos:
        1.250,50D -> 1250.50
        1.250,50C -> -1250.50
        vazio     -> 0.0
    """
    if pd.isna(valor) or str(valor).strip() == "":
        return 0.0

    texto = str(valor).strip().upper()
    sinal = 1

    if texto.endswith("D"):
        numero_str = texto[:-1].strip()
        sinal = 1

    elif texto.endswith("C"):
        numero_str = texto[:-1].strip()
        sinal = -1

    else:
        numero_str = texto

    try:
        numero_str = (
            numero_str
            .replace(".", "")
            .replace(",", ".")
        )

        return float(numero_str) * sinal

    except (ValueError, TypeError):
        return 0.0


def _converter_serie_decimal_brasileiro(serie):
    """
    Converte uma série do Pandas com números no formato brasileiro.

    Exemplos:
        388788112,3 -> 388788112.3
        1.250,50    -> 1250.50
        vazio       -> 0.0
    """
    texto = (
        serie
        .fillna("")
        .astype(str)
        .str.strip()
        .str.replace(".", "", regex=False)
        .str.replace(",", ".", regex=False)
    )

    return pd.to_numeric(
        texto,
        errors="coerce"
    ).fillna(0.0)


def _obter_engine_excel(caminho_arquivo):
    """
    Define o mecanismo de leitura conforme a extensão do arquivo.

    Arquivos .xls utilizam xlrd.
    Arquivos .xlsx utilizam openpyxl.
    """
    extensao = os.path.splitext(caminho_arquivo)[1].lower()

    if extensao == ".xls":
        return "xlrd"

    return "openpyxl"


def _ler_csv_useall(caminho_arquivo):
    """
    Lê um arquivo CSV exportado pelo Useall.

    O arquivo deve utilizar ponto e vírgula como separador.

    A função tenta as seguintes codificações:
        1. UTF-8 com possível BOM;
        2. Latin-1;
        3. Windows-1252.
    """
    nome_arquivo = os.path.basename(caminho_arquivo)

    configuracao_leitura = {
        "sep": ";",
        "dtype": str,
        "keep_default_na": False,
    }

    codificacoes = [
        "utf-8-sig",
        "latin-1",
        "cp1252",
    ]

    ultimo_erro = None

    for codificacao in codificacoes:
        try:
            return pd.read_csv(
                caminho_arquivo,
                encoding=codificacao,
                **configuracao_leitura
            )

        except UnicodeDecodeError as erro:
            ultimo_erro = erro

        except Exception as erro:
            raise ValueError(
                f"Não foi possível ler o arquivo CSV "
                f"'{nome_arquivo}'. Erro: {erro}"
            ) from erro

    raise ValueError(
        f"Não foi possível identificar a codificação do arquivo CSV "
        f"'{nome_arquivo}'. Erro: {ultimo_erro}"
    )


# ==============================================================================
# TRANSFORMAÇÃO COMUM DOS BALANCETES EM EXCEL
# ==============================================================================

def transformar_balancete(caminho_arquivo):
    """
    Transforma um balancete no layout comum utilizado pelos sistemas
    que trabalham com arquivos Excel.
    """
    nome_arquivo = os.path.basename(caminho_arquivo)

    try:
        df_origem = pd.read_excel(
            caminho_arquivo,
            sheet_name=0,
            dtype=str,
            engine=_obter_engine_excel(caminho_arquivo)
        )

        df_origem.dropna(
            how="all",
            inplace=True
        )

    except Exception as erro:
        raise ValueError(
            f"Não foi possível ler o arquivo {nome_arquivo}. "
            f"Erro: {erro}"
        ) from erro

    if df_origem.shape[1] < 6:
        raise ValueError(
            f"O arquivo '{nome_arquivo}' possui "
            f"{df_origem.shape[1]} coluna(s), mas são necessárias "
            f"pelo menos 6 colunas para transformar o balancete."
        )

    df_origem.reset_index(
        drop=True,
        inplace=True
    )

    df_destino = pd.DataFrame(
        index=df_origem.index
    )

    df_destino["Atividade"] = "Geral"

    df_destino["Conta"] = (
        df_origem.iloc[:, 0]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    df_destino["Nome"] = (
        df_origem.iloc[:, 1]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    df_destino["Cód. Reduzido"] = (
        df_origem.iloc[:, 0]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    saldo_anterior = pd.to_numeric(
        df_origem.iloc[:, 2]
        .fillna("")
        .astype(str)
        .str.replace(",", ".", regex=False),
        errors="coerce"
    ).fillna(0.0)

    debito = pd.to_numeric(
        df_origem.iloc[:, 3]
        .fillna("")
        .astype(str)
        .str.replace(",", ".", regex=False),
        errors="coerce"
    ).fillna(0.0)

    credito = pd.to_numeric(
        df_origem.iloc[:, 4]
        .fillna("")
        .astype(str)
        .str.replace(",", ".", regex=False),
        errors="coerce"
    ).fillna(0.0)

    saldo_acumulado_origem = pd.to_numeric(
        df_origem.iloc[:, 5]
        .fillna("")
        .astype(str)
        .str.replace(",", ".", regex=False),
        errors="coerce"
    ).fillna(0.0)

    df_destino["Saldo Anterior"] = -saldo_anterior
    df_destino["Débito"] = debito
    df_destino["Crédito"] = credito
    df_destino["Movimento"] = debito - credito
    df_destino["Saldo Acumulado"] = -saldo_acumulado_origem

    return df_destino


# ==============================================================================
# TRANSFORMAÇÃO DO PLANO DE CONTAS
# ==============================================================================

def transformar_plano_de_contas(caminho_arquivo):
    """
    Transforma um plano de contas no layout esperado pelo HUB.
    """
    nome_arquivo = os.path.basename(caminho_arquivo)

    try:
        df_origem = pd.read_excel(
            caminho_arquivo,
            sheet_name=0,
            skiprows=5,
            dtype=str,
            engine=_obter_engine_excel(caminho_arquivo)
        )

        df_origem.dropna(
            how="all",
            inplace=True
        )

    except Exception as erro:
        raise ValueError(
            f"Não foi possível ler o plano de contas "
            f"{nome_arquivo}. Erro: {erro}"
        ) from erro

    if df_origem.shape[1] < 4:
        raise ValueError(
            f"O plano de contas '{nome_arquivo}' possui "
            f"{df_origem.shape[1]} coluna(s), mas são necessárias "
            f"pelo menos 4 colunas para realizar a transformação."
        )

    df_origem.reset_index(
        drop=True,
        inplace=True
    )

    df_destino = pd.DataFrame(
        index=df_origem.index
    )

    classif = (
        df_origem.iloc[:, 2]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    df_destino["Chave Cliente"] = classif

    df_destino["Chave D&M"] = (
        classif
        .str[0]
        .where(
            classif.str[0].isin(["1", "2", "3"]),
            ""
        )
    )

    df_destino["Classificação"] = (
        df_origem.iloc[:, 0]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    df_destino["Descrição"] = (
        df_origem.iloc[:, 3]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    df_destino["Sint./An."] = (
        df_origem.iloc[:, 1]
        .fillna("")
        .astype(str)
        .str.strip()
        .str.upper()
        .apply(
            lambda valor: "S" if valor == "T" else "A"
        )
    )

    def definir_tipo_conta(classificacao):
        """
        Define o tipo da conta com base no primeiro caractere
        da classificação.
        """
        if (
            not isinstance(classificacao, str)
            or not classificacao
        ):
            return "R"

        primeiro_caractere = classificacao[0]

        if primeiro_caractere == "1":
            return "A"

        if primeiro_caractere == "2":
            return "P"

        return "R"

    df_destino["At/Pas/Res"] = classif.apply(
        definir_tipo_conta
    )

    df_destino["Indice"] = range(
        1,
        len(df_destino) + 1
    )

    return df_destino


# ==============================================================================
# TRANSFORMAÇÃO DO BALANCETE USEALL
# ==============================================================================

def transformar_balancete_useall(caminho_arquivo):
    """
    Transforma o balancete CSV exportado pelo sistema Useall.

    Layout final:
        Coluna A = Atividade
        Coluna B = Conta
        Coluna C = Nome
        Coluna D = Cód. Reduzido
        Coluna E = Saldo Anterior
        Coluna F = Débito
        Coluna G = Crédito
        Coluna H = Movimento
        Coluna I = Saldo Acumulado

    Mapeamento:
        Valor fixo "Geral"        -> Atividade
        Origem B - ClasMasc       -> Conta
        Origem C - NomeConta      -> Nome
        Origem A - CodigoConta    -> Cód. Reduzido
        Origem D - SaldoAnterior  -> Saldo Anterior
        Origem E - Debitos        -> Débito
        Origem F - Creditos       -> Crédito
        Débito menos Crédito      -> Movimento
        Origem G - SaldoFinal     -> Saldo Acumulado
    """
    nome_arquivo = os.path.basename(caminho_arquivo)
    extensao = os.path.splitext(caminho_arquivo)[1].lower()

    if extensao != ".csv":
        raise ValueError(
            f"O arquivo '{nome_arquivo}' não possui extensão CSV."
        )

    df_origem = _ler_csv_useall(caminho_arquivo)

    # Remove espaços e uma possível marca BOM dos cabeçalhos.
    df_origem.columns = [
        str(coluna)
        .replace("\ufeff", "")
        .strip()
        for coluna in df_origem.columns
    ]

    colunas_obrigatorias = [
        "CodigoConta",
        "ClasMasc",
        "NomeConta",
        "SaldoAnterior",
        "Debitos",
        "Creditos",
        "SaldoFinal",
    ]

    colunas_ausentes = [
        coluna
        for coluna in colunas_obrigatorias
        if coluna not in df_origem.columns
    ]

    if colunas_ausentes:
        colunas_encontradas = ", ".join(
            str(coluna)
            for coluna in df_origem.columns
        )

        raise ValueError(
            f"O arquivo '{nome_arquivo}' não possui todas as "
            f"colunas obrigatórias do Useall. "
            f"Colunas ausentes: {', '.join(colunas_ausentes)}. "
            f"Colunas encontradas: {colunas_encontradas}."
        )

    # Converte campos contendo somente espaços em valores ausentes.
    df_origem = df_origem.replace(
        r"^\s*$",
        pd.NA,
        regex=True
    )

    # Remove linhas completamente vazias nas colunas utilizadas.
    df_origem.dropna(
        how="all",
        subset=colunas_obrigatorias,
        inplace=True
    )

    df_origem.reset_index(
        drop=True,
        inplace=True
    )

    # Inicializa o DataFrame com o mesmo índice da origem.
    # Isso garante que o valor "Geral" seja repetido em todas as linhas.
    df_destino = pd.DataFrame(
        index=df_origem.index
    )

    # Coluna A: valor fixo
    df_destino["Atividade"] = "Geral"

    # Coluna B: origem B, ClasMasc
    df_destino["Conta"] = (
        df_origem["ClasMasc"]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    # Coluna C: origem C, NomeConta
    df_destino["Nome"] = (
        df_origem["NomeConta"]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    # Coluna D: origem A, CodigoConta
    df_destino["Cód. Reduzido"] = (
        df_origem["CodigoConta"]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    # Converte os campos monetários antes de montar as colunas.
    saldo_anterior = _converter_serie_decimal_brasileiro(
        df_origem["SaldoAnterior"]
    )

    debito = _converter_serie_decimal_brasileiro(
        df_origem["Debitos"]
    )

    credito = _converter_serie_decimal_brasileiro(
        df_origem["Creditos"]
    )

    saldo_acumulado = _converter_serie_decimal_brasileiro(
        df_origem["SaldoFinal"]
    )

    # Coluna E: origem D, SaldoAnterior
    df_destino["Saldo Anterior"] = saldo_anterior

    # Coluna F: origem E, Debitos
    df_destino["Débito"] = debito

    # Coluna G: origem F, Creditos
    df_destino["Crédito"] = credito

    # Coluna H: Débito menos Crédito
    df_destino["Movimento"] = debito - credito

    # Coluna I: origem G, SaldoFinal
    df_destino["Saldo Acumulado"] = saldo_acumulado

    return df_destino

def _converter_numero_cooperoque(valor):
    """
    Converte valores do balancete Cooperoque para número.

    Aceita:
    - Valores numéricos provenientes diretamente do Excel;
    - Formato brasileiro: 263.167.952,92;
    - Formato decimal: 263167952.92;
    - Células vazias.

    Exemplos:
        263.167.952,92 -> 263167952.92
        263167952.92   -> 263167952.92
        vazio          -> 0.0
    """
    if pd.isna(valor):
        return 0.0

    if isinstance(valor, (int, float)):
        return float(valor)

    texto = str(valor).strip()

    if texto == "":
        return 0.0

    # Remove espaços comuns e espaços não separáveis.
    texto = (
        texto
        .replace("\xa0", "")
        .replace(" ", "")
    )

    # Remove possíveis símbolos monetários.
    texto = (
        texto
        .replace("R$", "")
        .replace("$", "")
    )

    # Trata valores negativos entre parênteses.
    negativo_por_parenteses = (
        texto.startswith("(")
        and texto.endswith(")")
    )

    if negativo_por_parenteses:
        texto = texto[1:-1]

    # Se existir vírgula, considera o formato brasileiro.
    if "," in texto:
        texto = (
            texto
            .replace(".", "")
            .replace(",", ".")
        )

    # Se não houver vírgula, mantém o ponto como separador decimal.
    try:
        numero = float(texto)

        if negativo_por_parenteses:
            numero = -abs(numero)

        return numero

    except (ValueError, TypeError):
        return 0.0


def _aplicar_natureza_cooperoque(valor, natureza):
    """
    Aplica o sinal do saldo conforme a natureza contábil.

    Regras:
        D = positivo
        C = negativo

    Se a natureza estiver vazia ou for desconhecida, mantém o sinal
    original do valor.
    """
    numero = _converter_numero_cooperoque(valor)

    if pd.isna(natureza):
        return numero

    natureza_texto = str(natureza).strip().upper()

    if natureza_texto == "D":
        return abs(numero)

    if natureza_texto == "C":
        return -abs(numero)

    return numero


def transformar_balancete_cooperoque(caminho_arquivo):
    """
    Transforma o balancete Excel do cliente Cooperoque.

    Layout de origem:
        A = Código da Conta
        B = Descrição
        C = Saldo Anterior
        D = Natureza do Saldo Anterior, D ou C
        E = Débitos
        F = Créditos
        G = Saldo Atual
        H = Natureza do Saldo Atual, D ou C

    Layout de destino:
        A = Atividade
        B = Conta
        C = Nome
        D = Cód. Reduzido
        E = Saldo Anterior
        F = Débito
        G = Crédito
        H = Movimento
        I = Saldo Acumulado
    """
    nome_arquivo = os.path.basename(caminho_arquivo)
    extensao = os.path.splitext(caminho_arquivo)[1].lower()

    if extensao not in {".xls", ".xlsx"}:
        raise ValueError(
            f"O arquivo '{nome_arquivo}' não é um arquivo Excel válido."
        )

    engine = "xlrd" if extensao == ".xls" else "openpyxl"

    try:
        df_origem = pd.read_excel(
            caminho_arquivo,
            sheet_name=0,
            dtype=object,
            engine=engine
        )

    except Exception as erro:
        raise ValueError(
            f"Não foi possível ler o arquivo"
            f"'{nome_arquivo}'. Erro: {erro}"
        ) from erro

    if df_origem.shape[1] < 8:
        raise ValueError(
            f"O arquivo '{nome_arquivo}' possui "
            f"{df_origem.shape[1]} coluna(s), mas são esperadas "
            "pelo menos 8 colunas, de A até H."
        )

    # Remove linhas completamente vazias.
    df_origem.dropna(
        how="all",
        inplace=True
    )

    # Remove linhas sem código de conta e sem descrição.
    linhas_validas = (
        df_origem.iloc[:, 0].notna()
        | df_origem.iloc[:, 1].notna()
    )

    df_origem = df_origem.loc[
        linhas_validas
    ].copy()

    df_origem.reset_index(
        drop=True,
        inplace=True
    )

    if df_origem.empty:
        raise ValueError(
            f"O arquivo '{nome_arquivo}' não possui "
            "linhas válidas para tabulação."
        )

    df_destino = pd.DataFrame(
        index=df_origem.index
    )

    # Coluna A: valor fixo.
    df_destino["Atividade"] = "Geral"

    # Coluna B: origem A, Código da Conta.
    df_destino["Conta"] = (
        df_origem.iloc[:, 0]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    # Coluna C: origem B, Descrição.
    df_destino["Nome"] = (
        df_origem.iloc[:, 1]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    # Coluna D: origem A, Código da Conta.
    df_destino["Cód. Reduzido"] = (
        df_origem.iloc[:, 0]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    # Coluna E: origem C com natureza definida pela origem D.
    saldo_anterior = pd.Series(
        [
            _aplicar_natureza_cooperoque(valor, natureza)
            for valor, natureza in zip(
                df_origem.iloc[:, 2],
                df_origem.iloc[:, 3]
            )
        ],
        index=df_origem.index,
        dtype=float
    )

    # Coluna F: origem E.
    debito = df_origem.iloc[:, 4].apply(
        _converter_numero_cooperoque
    )

    # Coluna G: origem F.
    credito = df_origem.iloc[:, 5].apply(
        _converter_numero_cooperoque
    )

    # Coluna I: origem G com natureza definida pela origem H.
    saldo_acumulado = pd.Series(
        [
            _aplicar_natureza_cooperoque(valor, natureza)
            for valor, natureza in zip(
                df_origem.iloc[:, 6],
                df_origem.iloc[:, 7]
            )
        ],
        index=df_origem.index,
        dtype=float
    )

    df_destino["Saldo Anterior"] = saldo_anterior
    df_destino["Débito"] = debito
    df_destino["Crédito"] = credito

    # Coluna H.
    df_destino["Movimento"] = debito - credito

    # Coluna I.
    df_destino["Saldo Acumulado"] = saldo_acumulado

    return df_destino

def _normalizar_texto_coopercargo(valor):
    """
    Normaliza um texto para comparação.

    Remove espaços extras, espaços não separáveis e converte
    o conteúdo para letras maiúsculas.
    """
    if pd.isna(valor):
        return ""

    return (
        str(valor)
        .replace("\xa0", " ")
        .strip()
        .upper()
    )


def _converter_numero_coopercargo(valor):
    """
    Converte valores monetários do cliente Coopercargo para float.

    Formatos aceitos:
        92.045.259,67
        92045259,67
        92045259.67
        92045259
        (1.250,50)
        R$ 1.250,50

    Valores vazios ou inválidos são convertidos para 0.0.
    """
    if pd.isna(valor):
        return 0.0

    if isinstance(valor, (int, float)):
        return float(valor)

    texto = str(valor).strip()

    if not texto:
        return 0.0

    texto = (
        texto
        .replace("\xa0", "")
        .replace(" ", "")
        .replace("R$", "")
        .replace("$", "")
    )

    negativo_por_parenteses = (
        texto.startswith("(")
        and texto.endswith(")")
    )

    if negativo_por_parenteses:
        texto = texto[1:-1]

    # Quando há vírgula, considera o padrão brasileiro.
    if "," in texto:
        texto = (
            texto
            .replace(".", "")
            .replace(",", ".")
        )

    try:
        numero = float(texto)

        if negativo_por_parenteses:
            numero = -abs(numero)

        return numero

    except (ValueError, TypeError):
        return 0.0


def _aplicar_natureza_coopercargo(valor, natureza):
    """
    Aplica o sinal conforme a natureza contábil.

    Regras:
        D = positivo
        C = negativo

    Caso a natureza esteja vazia ou seja desconhecida,
    mantém o sinal original do valor.
    """
    numero = _converter_numero_coopercargo(valor)
    natureza_texto = _normalizar_texto_coopercargo(natureza)

    if natureza_texto == "D":
        return abs(numero)

    if natureza_texto == "C":
        return -abs(numero)

    return numero


def _localizar_inicio_balancete_coopercargo(df_origem, nome_arquivo):
    """
    Localiza a primeira linha cuja descrição, na coluna D,
    seja exatamente ATIVO.

    Todas as linhas anteriores são desconsideradas.
    A própria linha ATIVO é mantida no resultado.
    """
    if df_origem.shape[1] < 4:
        raise ValueError(
            f"O arquivo '{nome_arquivo}' não possui a coluna D, "
            "necessária para localizar a descrição ATIVO."
        )

    descricoes = (
        df_origem.iloc[:, 3]
        .apply(_normalizar_texto_coopercargo)
    )

    indices_ativo = df_origem.index[
        descricoes == "ATIVO"
    ]

    if len(indices_ativo) == 0:
        raise ValueError(
            f"Não foi encontrada uma conta com a descrição "
            f"'ATIVO' na coluna D do arquivo '{nome_arquivo}'."
        )

    return indices_ativo[0]


def transformar_balancete_coopercargo(caminho_arquivo):
    """
    Transforma o balancete Excel do cliente Coopercargo.

    O arquivo é lido sem cabeçalho porque contém informações da
    empresa antes do balancete. O processamento começa na primeira
    linha cuja coluna D tenha a descrição ATIVO.

    Layout de origem:
        C = Classificação
        D = Descrição
        E = Red.
        F = Saldo Anterior
        G = Natureza do Saldo Anterior
        H = Débitos
        I = Créditos
        J = Movimento
        K = Natureza do Movimento
        L = Saldo Atual
        M = Natureza do Saldo Atual

    Layout de destino:
        A = Atividade
        B = Conta
        C = Nome
        D = Cód. Reduzido
        E = Saldo Anterior
        F = Débito
        G = Crédito
        H = Movimento
        I = Saldo Acumulado
    """
    nome_arquivo = os.path.basename(caminho_arquivo)
    extensao = os.path.splitext(caminho_arquivo)[1].lower()

    if extensao not in {".xls", ".xlsx"}:
        raise ValueError(
            f"O arquivo '{nome_arquivo}' não é um arquivo Excel válido."
        )

    engine = "xlrd" if extensao == ".xls" else "openpyxl"

    try:
        df_origem = pd.read_excel(
            caminho_arquivo,
            sheet_name=0,
            header=None,
            dtype=object,
            engine=engine
        )

    except Exception as erro:
        raise ValueError(
            f"Não foi possível ler o arquivo do cliente Coopercargo "
            f"'{nome_arquivo}'. Erro: {erro}"
        ) from erro

    if df_origem.shape[1] < 13:
        raise ValueError(
            f"O arquivo '{nome_arquivo}' possui "
            f"{df_origem.shape[1]} coluna(s), mas são necessárias "
            "pelo menos 13 colunas, de A até M."
        )

    indice_inicio = _localizar_inicio_balancete_coopercargo(
        df_origem,
        nome_arquivo
    )

    # Mantém a linha ATIVO e todas as linhas posteriores.
    df_origem = df_origem.loc[
        indice_inicio:
    ].copy()

    # Remove linhas completamente vazias.
    df_origem.dropna(
        how="all",
        inplace=True
    )

    # Mantém linhas que tenham classificação ou descrição.
    classificacao_preenchida = (
        df_origem.iloc[:, 2]
        .fillna("")
        .astype(str)
        .str.strip()
        .ne("")
    )

    descricao_preenchida = (
        df_origem.iloc[:, 3]
        .fillna("")
        .astype(str)
        .str.strip()
        .ne("")
    )

    df_origem = df_origem.loc[
        classificacao_preenchida
        | descricao_preenchida
    ].copy()

    df_origem.reset_index(
        drop=True,
        inplace=True
    )

    if df_origem.empty:
        raise ValueError(
            f"O arquivo '{nome_arquivo}' não possui linhas válidas "
            "após a conta ATIVO."
        )

    df_destino = pd.DataFrame(
        index=df_origem.index
    )

    # Coluna A: valor fixo.
    df_destino["Atividade"] = "Geral"

    # Coluna B: origem C, Classificação.
    df_destino["Conta"] = (
        df_origem.iloc[:, 2]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    # Coluna C: origem D, Descrição.
    df_destino["Nome"] = (
        df_origem.iloc[:, 3]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    # Coluna D: origem E, Código reduzido.
    df_destino["Cód. Reduzido"] = (
        df_origem.iloc[:, 4]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    # Coluna E: origem F, sinal definido pela origem G.
    saldo_anterior = pd.Series(
        [
            _aplicar_natureza_coopercargo(valor, natureza)
            for valor, natureza in zip(
                df_origem.iloc[:, 5],
                df_origem.iloc[:, 6]
            )
        ],
        index=df_origem.index,
        dtype=float
    )

    # Coluna F: origem H.
    debito = (
        df_origem.iloc[:, 7]
        .apply(_converter_numero_coopercargo)
        .astype(float)
    )

    # Coluna G: origem I.
    credito = (
        df_origem.iloc[:, 8]
        .apply(_converter_numero_coopercargo)
        .astype(float)
    )

    # Coluna H: origem J com sinal definido pela origem K.
    movimento = pd.Series(
        [
            _aplicar_natureza_coopercargo(valor, natureza)
            for valor, natureza in zip(
                df_origem.iloc[:, 9],
                df_origem.iloc[:, 10]
            )
        ],
        index=df_origem.index,
        dtype=float
    )

    # Coluna I: origem L com sinal definido pela origem M.
    saldo_acumulado = pd.Series(
        [
            _aplicar_natureza_coopercargo(valor, natureza)
            for valor, natureza in zip(
                df_origem.iloc[:, 11],
                df_origem.iloc[:, 12]
            )
        ],
        index=df_origem.index,
        dtype=float
    )

    df_destino["Saldo Anterior"] = saldo_anterior
    df_destino["Débito"] = debito
    df_destino["Crédito"] = credito
    df_destino["Movimento"] = movimento
    df_destino["Saldo Acumulado"] = saldo_acumulado

    return df_destino

# ==============================================================================
# TRANSFORMAÇÃO DO BALANCETE COAGRIL
# ==============================================================================

def _normalizar_texto_coagril(valor):
    """
    Normaliza textos utilizados nas comparações do Cliente Coagril.

    A função:
    - substitui espaços não separáveis;
    - remove espaços no início e no final;
    - converte o texto para letras maiúsculas.
    """
    if valor is None:
        return ""

    return (
        str(valor)
        .replace("\xa0", " ")
        .strip()
        .upper()
    )


def _converter_numero_coagril(valor):
    """
    Converte valores monetários do TXT para float.

    Formatos aceitos:
        476.164.985,07
        -203.558.288,59
        92045259,67
        92045259.67
        0,00
        (1.250,50)

    Valores vazios ou inválidos são convertidos para 0.0.
    """
    if valor is None:
        return 0.0

    if isinstance(valor, (int, float)):
        return float(valor)

    texto = str(valor).strip()

    if not texto:
        return 0.0

    texto = (
        texto
        .replace("\xa0", "")
        .replace(" ", "")
        .replace("R$", "")
        .replace("$", "")
    )

    negativo_por_parenteses = (
        texto.startswith("(")
        and texto.endswith(")")
    )

    if negativo_por_parenteses:
        texto = texto[1:-1]

    # Quando existe vírgula, utiliza o formato monetário brasileiro.
    if "," in texto:
        texto = (
            texto
            .replace(".", "")
            .replace(",", ".")
        )

    try:
        numero = float(texto)

        if negativo_por_parenteses:
            numero = -abs(numero)

        return numero

    except (ValueError, TypeError):
        return 0.0


def _ler_linhas_txt_coagril(caminho_arquivo):
    """
    Lê as linhas do arquivo TXT tentando diferentes codificações.

    Ordem das codificações:
        1. UTF-8 com possível BOM;
        2. Windows-1252;
        3. Latin-1.
    """
    nome_arquivo = os.path.basename(caminho_arquivo)

    codificacoes = [
        "utf-8-sig",
        "cp1252",
        "latin-1",
    ]

    ultimo_erro = None

    for codificacao in codificacoes:
        try:
            with open(
                caminho_arquivo,
                mode="r",
                encoding=codificacao
            ) as arquivo:
                return arquivo.readlines()

        except UnicodeDecodeError as erro:
            ultimo_erro = erro

        except OSError as erro:
            raise ValueError(
                f"Não foi possível abrir o arquivo TXT "
                f"'{nome_arquivo}'. Erro: {erro}"
            ) from erro

    raise ValueError(
        f"Não foi possível identificar a codificação do arquivo TXT "
        f"'{nome_arquivo}'. Erro: {ultimo_erro}"
    )


def _extrair_linha_contabil_coagril(linha):
    """
    Extrai uma linha contábil válida do TXT do cliente Coagril.

    A conta pode ser composta por vários números separados por espaços.

    Exemplos:
        1                   ATIVO
        1 1                 ATIVO CIRCULANTE
        1 1 1               DISPONIBILIDADES
        1 1 1 01            CAIXA GERAL

    Estrutura esperada:
        Conta
        Nomenclatura
        Saldo Anterior
        Débito
        Crédito
        Saldo Atual

    Retorna um dicionário quando a linha for contábil.
    Retorna None para cabeçalhos, informações da empresa e linhas vazias.
    """
    if linha is None:
        return None

    linha = (
        str(linha)
        .replace("\xa0", " ")
        .replace("\r", "")
        .replace("\n", "")
        .rstrip()
    )

    if not linha.strip():
        return None

    # Formato monetário brasileiro com sinal opcional.
    #
    # Exemplos reconhecidos:
    # 476.164.985,07
    # -203.558.288,59
    # 0,00
    formato_monetario = (
        r"-?"
        r"(?:\d{1,3}(?:\.\d{3})*|\d+)"
        r",\d{2}"
    )

    # A conta pode possuir vários níveis numéricos separados por espaços.
    #
    # Exemplos:
    # 1
    # 1 1
    # 1 1 1
    # 1 1 1 01
    #
    # A separação entre conta e descrição deve possuir pelo menos
    # dois espaços, conforme o layout fixo do relatório.
    padrao_linha = re.compile(
        r"^\s*"
        r"(?P<conta>\d+(?:[ \t]+\d+)*)"
        r"[ \t]{2,}"
        r"(?P<nome>.+?\S)"
        r"[ \t]{2,}"
        rf"(?P<saldo_anterior>{formato_monetario})"
        r"[ \t]+"
        rf"(?P<debito>{formato_monetario})"
        r"[ \t]+"
        rf"(?P<credito>{formato_monetario})"
        r"[ \t]+"
        rf"(?P<saldo_atual>{formato_monetario})"
        r"\s*$",
        re.IGNORECASE
    )

    correspondencia = padrao_linha.match(linha)

    if correspondencia is None:
        return None

    # Preserva todos os níveis da conta, mas padroniza a quantidade
    # de espaços entre eles para apenas um espaço.
    conta = re.sub(
        r"\s+",
        " ",
        correspondencia.group("conta")
    ).strip()

    # Remove espaços excedentes da descrição.
    nome = re.sub(
        r"\s+",
        " ",
        correspondencia.group("nome")
    ).strip()

    if not conta or not nome:
        return None

    # Proteção adicional contra possíveis cabeçalhos.
    nomes_invalidos = {
        "NOMENCLATURA",
        "N O M E N C L A T U R A",
        "DESCRIÇÃO",
        "DESCRICAO",
        "NOME",
    }

    if nome.upper() in nomes_invalidos:
        return None

    return {
        "Conta": conta,
        "Nome": nome,
        "Saldo Anterior": correspondencia.group(
            "saldo_anterior"
        ),
        "Débito": correspondencia.group(
            "debito"
        ),
        "Crédito": correspondencia.group(
            "credito"
        ),
        "Saldo Acumulado": correspondencia.group(
            "saldo_atual"
        ),
    }


def _localizar_inicio_coagril(registros, nome_arquivo):
    """
    Localiza o primeiro registro cuja descrição seja ATIVO.

    Todos os registros contábeis anteriores são desconsiderados.
    A própria linha ATIVO é mantida.
    """
    for indice, registro in enumerate(registros):
        descricao = _normalizar_texto_coagril(
            registro["Nome"]
        )

        if descricao == "ATIVO":
            return indice

    raise ValueError(
        f"Não foi encontrada uma conta com a descrição 'ATIVO' "
        f"no arquivo '{nome_arquivo}'."
    )


def transformar_balancete_coagril(caminho_arquivo):
    """
    Transforma o balancete TXT do Cliente Coagril.

    O processamento começa na primeira linha contábil cuja descrição
    seja ATIVO.

    Cabeçalhos, informações da empresa e linhas intermediárias que não
    seguirem o formato contábil são automaticamente desconsiderados.

    Layout de origem:
        A = Conta
        B = Nomenclatura
        C = Saldo Anterior
        D = Débito
        E = Crédito
        F = Saldo Atual

    Layout de destino:
        A = Atividade
        B = Conta
        C = Nome
        D = Cód. Reduzido
        E = Saldo Anterior
        F = Débito
        G = Crédito
        H = Movimento
        I = Saldo Acumulado

    Cálculo:
        Movimento = Débito + Crédito

    O crédito é somado porque o arquivo de origem já apresenta
    os créditos com sinal negativo.
    """
    nome_arquivo = os.path.basename(caminho_arquivo)
    extensao = os.path.splitext(caminho_arquivo)[1].lower()

    if extensao != ".txt":
        raise ValueError(
            f"O arquivo '{nome_arquivo}' não possui extensão TXT."
        )

    linhas = _ler_linhas_txt_coagril(
        caminho_arquivo
    )

    if not linhas:
        raise ValueError(
            f"O arquivo TXT '{nome_arquivo}' está vazio."
        )

    registros = []

    for linha in linhas:
        registro = _extrair_linha_contabil_coagril(
            linha
        )

        if registro is not None:
            registros.append(registro)

    if not registros:
        raise ValueError(
            f"Nenhuma linha contábil válida foi encontrada no "
            f"arquivo '{nome_arquivo}'."
        )

    indice_inicio = _localizar_inicio_coagril(
        registros,
        nome_arquivo
    )

    # Desconsidera todos os registros anteriores à primeira conta ATIVO.
    registros = registros[indice_inicio:]

    if not registros:
        raise ValueError(
            f"O arquivo '{nome_arquivo}' não possui registros "
            "contábeis após a conta ATIVO."
        )

    df_origem = pd.DataFrame(registros)

    df_destino = pd.DataFrame(
        index=df_origem.index
    )

    # Coluna A: valor fixo.
    df_destino["Atividade"] = "Geral"

    # Coluna B: origem A.
    df_destino["Conta"] = (
        df_origem["Conta"]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    # Coluna C: origem B.
    df_destino["Nome"] = (
        df_origem["Nome"]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    # Coluna D: origem A.
    df_destino["Cód. Reduzido"] = (
        df_origem["Conta"]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    # Coluna E: origem C.
    saldo_anterior = (
        df_origem["Saldo Anterior"]
        .apply(_converter_numero_coagril)
        .astype(float)
    )

    # Coluna F: origem D.
    debito = (
        df_origem["Débito"]
        .apply(_converter_numero_coagril)
        .astype(float)
    )

    # Coluna G: origem E.
    credito = (
        df_origem["Crédito"]
        .apply(_converter_numero_coagril)
        .astype(float)
    )

    # Coluna I: origem F.
    saldo_acumulado = (
        df_origem["Saldo Acumulado"]
        .apply(_converter_numero_coagril)
        .astype(float)
    )

    df_destino["Saldo Anterior"] = saldo_anterior
    df_destino["Débito"] = debito
    df_destino["Crédito"] = credito

    # Coluna H.
    # O crédito já vem negativo no arquivo original.
    df_destino["Movimento"] = debito + credito

    # Coluna I.
    df_destino["Saldo Acumulado"] = saldo_acumulado

    return df_destino

# ==============================================================================
# TRANSFORMAÇÃO DO BALANCETE Fecoagro SC
# ==============================================================================

def _normalizar_classificacao_fecoagrosc(valor):
    """
    Normaliza a classificação contábil do cliente Fecoagro SC.

    Todas as classificações possuem 15 dígitos e utilizam zeros à
    direita para completar o tamanho.

    A quantidade de dígitos significativos é determinada pelos níveis
    hierárquicos válidos da classificação:

        Nível 1:  1 dígito
        Nível 2:  3 dígitos
        Nível 3:  5 dígitos
        Nível 4:  7 dígitos
        Nível 5:  9 dígitos
        Nível 6: 11 dígitos
        Nível 7: 13 dígitos
        Nível 8: 15 dígitos

    Dessa forma, um zero localizado dentro do nível da conta será
    preservado, mesmo que esteja no final da classificação real.

    Exemplos:
        100000000000000 -> 1
        101000000000000 -> 101
        101010000000000 -> 10101
        101010310000000 -> 101010310
        101010310100000 -> 1010103101? Não. Será enquadrado em 11
                           dígitos: 10101031010
    """
    if pd.isna(valor):
        return ""

    texto = str(valor).strip()

    if not texto:
        return ""

    # Remove espaços comuns e espaços não separáveis.
    texto = (
        texto
        .replace("\xa0", "")
        .replace(" ", "")
    )

    # Corrige valores eventualmente lidos como:
    # 100000000000000.0
    if re.fullmatch(r"\d+\.0+", texto):
        texto = texto.split(".", maxsplit=1)[0]

    # Corrige valores eventualmente apresentados em notação científica.
    if re.fullmatch(
        r"[+-]?\d+(?:[.,]\d+)?[Ee][+-]?\d+",
        texto
    ):
        try:
            from decimal import Decimal, InvalidOperation

            texto_decimal = texto.replace(",", ".")

            texto = format(
                Decimal(texto_decimal),
                "f"
            )

            if "." in texto:
                parte_inteira, parte_decimal = texto.split(
                    ".",
                    maxsplit=1
                )

                if set(parte_decimal) <= {"0"}:
                    texto = parte_inteira

        except (InvalidOperation, ValueError):
            return str(valor).strip()

    # Se a classificação contiver outros caracteres, preserva o valor.
    if not texto.isdigit():
        return texto

    # Garante o padrão de 15 posições utilizado pelo sistema.
    texto = texto.zfill(15)

    niveis_validos = [
        1,
        3,
        5,
        7,
        9,
        11,
        13,
        15,
    ]

    # Localiza a posição do último dígito diferente de zero.
    ultima_posicao_significativa = 0

    for indice, caractere in enumerate(texto, start=1):
        if caractere != "0":
            ultima_posicao_significativa = indice

    # Se a classificação for composta somente por zeros.
    if ultima_posicao_significativa == 0:
        return "0"

    # Seleciona o primeiro nível hierárquico capaz de contemplar
    # o último dígito diferente de zero.
    tamanho_classificacao = 15

    for nivel in niveis_validos:
        if nivel >= ultima_posicao_significativa:
            tamanho_classificacao = nivel
            break

    return texto[:tamanho_classificacao]


def _converter_numero_fecoagrosc(valor):
    """
    Converte os valores monetários do cliente Fecoagro SC para float.

    Formatos aceitos:
        925.814.770,18
        925814770,18
        925814770.18
        -925.814.770,18
        (925.814.770,18)
        R$ 925.814.770,18

    Valores vazios ou inválidos são convertidos para 0.0.
    """
    if pd.isna(valor):
        return 0.0

    if isinstance(valor, (int, float)):
        return float(valor)

    texto = str(valor).strip()

    if not texto:
        return 0.0

    texto = (
        texto
        .replace("\xa0", "")
        .replace(" ", "")
        .replace("R$", "")
        .replace("$", "")
    )

    negativo_por_parenteses = (
        texto.startswith("(")
        and texto.endswith(")")
    )

    if negativo_por_parenteses:
        texto = texto[1:-1]

    # Se houver vírgula, considera o padrão monetário brasileiro.
    if "," in texto:
        texto = (
            texto
            .replace(".", "")
            .replace(",", ".")
        )

    try:
        numero = float(texto)

        if negativo_por_parenteses:
            numero = -abs(numero)

        return numero

    except (ValueError, TypeError):
        return 0.0


def transformar_balancete_fecoagrosc(caminho_arquivo):
    """
    Transforma o balancete Excel do cliente Fecoagro SC.

    Layout de origem:
        A = Ano
        B = Mês
        C = Conta
        D = ContaR
        E = Nome da Conta
        F = Filial
        G = Ativ
        H = CC
        I = Saldo Ant.
        J = Débito
        K = Crédito
        L = Saldo Mês
        M = Saldo Atual

    Layout de destino:
        A = Atividade
        B = Conta
        C = Nome
        D = Cód. Reduzido
        E = Saldo Anterior
        F = Débito
        G = Crédito
        H = Movimento
        I = Saldo Acumulado

    Mapeamento:
        Valor fixo "Geral"     -> Atividade
        Origem C               -> Conta
        Origem E               -> Nome
        Origem C               -> Cód. Reduzido
        Origem I               -> Saldo Anterior
        Origem J               -> Débito
        Origem K               -> Crédito
        Débito menos Crédito   -> Movimento
        Origem M               -> Saldo Acumulado
    """
    nome_arquivo = os.path.basename(caminho_arquivo)
    extensao = os.path.splitext(caminho_arquivo)[1].lower()

    if extensao not in {".xls", ".xlsx"}:
        raise ValueError(
            f"O arquivo '{nome_arquivo}' não é um arquivo Excel válido."
        )

    engine = "xlrd" if extensao == ".xls" else "openpyxl"

    try:
        df_origem = pd.read_excel(
            caminho_arquivo,
            sheet_name=0,
            dtype=str,
            engine=engine
        )

    except Exception as erro:
        raise ValueError(
            f"Não foi possível ler o arquivo do cliente Fecoagro SC "
            f"'{nome_arquivo}'. Erro: {erro}"
        ) from erro

    # A transformação utiliza até a coluna M.
    if df_origem.shape[1] < 13:
        raise ValueError(
            f"O arquivo do cliente Fecoagro SC '{nome_arquivo}' possui "
            f"{df_origem.shape[1]} coluna(s), mas são necessárias "
            "pelo menos 13 colunas, de A até M."
        )

    # Remove linhas completamente vazias.
    df_origem.dropna(
        how="all",
        inplace=True
    )

    # Mantém apenas linhas que tenham conta ou nome da conta.
    conta_preenchida = (
        df_origem.iloc[:, 2]
        .fillna("")
        .astype(str)
        .str.strip()
        .ne("")
    )

    nome_preenchido = (
        df_origem.iloc[:, 4]
        .fillna("")
        .astype(str)
        .str.strip()
        .ne("")
    )

    df_origem = df_origem.loc[
        conta_preenchida | nome_preenchido
    ].copy()

    df_origem.reset_index(
        drop=True,
        inplace=True
    )

    if df_origem.empty:
        raise ValueError(
            f"O arquivo do cliente Fecoagro SC '{nome_arquivo}' não possui "
            "linhas válidas para tabulação."
        )

    classificacao = (
        df_origem.iloc[:, 2]
        .apply(_normalizar_classificacao_fecoagrosc)
    )

    nome_conta = (
        df_origem.iloc[:, 4]
        .fillna("")
        .astype(str)
        .str.replace("\xa0", " ", regex=False)
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
    )

    saldo_anterior = (
        df_origem.iloc[:, 8]
        .apply(_converter_numero_fecoagrosc)
        .astype(float)
    )

    debito = (
        df_origem.iloc[:, 9]
        .apply(_converter_numero_fecoagrosc)
        .astype(float)
    )

    credito = (
        df_origem.iloc[:, 10]
        .apply(_converter_numero_fecoagrosc)
        .astype(float)
    )

    saldo_acumulado = (
        df_origem.iloc[:, 12]
        .apply(_converter_numero_fecoagrosc)
        .astype(float)
    )

    df_destino = pd.DataFrame(
        index=df_origem.index
    )

    # Coluna A.
    df_destino["Atividade"] = "Geral"

    # Coluna B: origem C normalizada.
    df_destino["Conta"] = classificacao

    # Coluna C: origem E.
    df_destino["Nome"] = nome_conta

    # Coluna D: origem C normalizada.
    df_destino["Cód. Reduzido"] = classificacao

    # Coluna E: origem I.
    df_destino["Saldo Anterior"] = saldo_anterior

    # Coluna F: origem J.
    df_destino["Débito"] = debito

    # Coluna G: origem K.
    df_destino["Crédito"] = credito

    # Coluna H: cálculo do movimento.
    df_destino["Movimento"] = debito - credito

    # Coluna I: origem M.
    df_destino["Saldo Acumulado"] = saldo_acumulado

    return df_destino
# ==============================================================================
# TRANSFORMAÇÃO DO BALANCETE COOPERLATE
# ==============================================================================

def _normalizar_texto_cooperlate(valor):
    """
    Normaliza textos utilizados no processamento do cliente Cooperlate.

    Remove espaços especiais, espaços repetidos e espaços nas
    extremidades.
    """
    if pd.isna(valor):
        return ""

    return re.sub(
        r"\s+",
        " ",
        str(valor).replace("\xa0", " ")
    ).strip()


def _converter_numero_cooperlate(valor):
    """
    Converte valores monetários do cliente Cooperlate para float.

    Formatos aceitos:
        145.430.616,65
        145430616,65
        145430616.65
        -145.430.616,65
        (145.430.616,65)
        R$ 145.430.616,65

    Valores vazios ou inválidos são convertidos para 0.0.
    """
    if pd.isna(valor):
        return 0.0

    if isinstance(valor, (int, float)):
        return float(valor)

    texto = str(valor).strip()

    if not texto:
        return 0.0

    texto = (
        texto
        .replace("\xa0", "")
        .replace(" ", "")
        .replace("R$", "")
        .replace("$", "")
    )

    negativo_por_parenteses = (
        texto.startswith("(")
        and texto.endswith(")")
    )

    if negativo_por_parenteses:
        texto = texto[1:-1]

    if "," in texto:
        texto = (
            texto
            .replace(".", "")
            .replace(",", ".")
        )

    try:
        numero = float(texto)

        if negativo_por_parenteses:
            numero = -abs(numero)

        return numero

    except (ValueError, TypeError):
        return 0.0


def _aplicar_natureza_cooperlate(valor, natureza):
    """
    Aplica a natureza contábil ao valor.

    Regras:
        D = positivo
        C = negativo

    Naturezas vazias ou desconhecidas mantêm o sinal original.
    """
    numero = _converter_numero_cooperlate(valor)

    natureza_normalizada = (
        _normalizar_texto_cooperlate(natureza)
        .upper()
    )

    if natureza_normalizada == "D":
        return abs(numero)

    if natureza_normalizada == "C":
        return -abs(numero)

    return numero


def _separar_conta_descricao_cooperlate(valor):
    """
    Separa a classificação e a descrição existentes na coluna A.

    Exemplos:
        "1 ATIVO"
            -> classificação: "1"
            -> descrição: "ATIVO"

        "101 ATIVO CIRCULANTE"
            -> classificação: "101"
            -> descrição: "ATIVO CIRCULANTE"

        "10101 DISPONIVEL"
            -> classificação: "10101"
            -> descrição: "DISPONIVEL"

    A classificação deve estar no início da célula e ser seguida
    por pelo menos um espaço antes da descrição.
    """
    texto = _normalizar_texto_cooperlate(valor)

    if not texto:
        return "", ""

    correspondencia = re.match(
        r"^(?P<conta>\d+)\s+(?P<descricao>.+?)$",
        texto
    )

    if correspondencia is None:
        return "", ""

    conta = correspondencia.group("conta").strip()
    descricao = correspondencia.group("descricao").strip()

    return conta, descricao


def _localizar_cabecalho_cooperlate(df_origem, nome_arquivo):
    """
    Localiza a linha de cabeçalho do balancete.

    O cabeçalho deve apresentar:
        Coluna A = Conta
        Coluna B = Cta red

    As linhas anteriores, como a indicação do período, são ignoradas.
    """
    for indice in df_origem.index:
        valor_coluna_a = (
            _normalizar_texto_cooperlate(
                df_origem.iloc[indice, 0]
            )
            .upper()
        )

        valor_coluna_b = (
            _normalizar_texto_cooperlate(
                df_origem.iloc[indice, 1]
            )
            .upper()
        )

        coluna_a_valida = valor_coluna_a == "CONTA"

        coluna_b_valida = valor_coluna_b in {
            "CTA RED",
            "CTA. RED",
            "CTA RED.",
            "CTA. RED.",
        }

        if coluna_a_valida and coluna_b_valida:
            return indice

    raise ValueError(
        f"Não foi possível localizar o cabeçalho com as colunas "
        f"'Conta' e 'Cta red' no arquivo '{nome_arquivo}'."
    )


def transformar_balancete_cooperlate(caminho_arquivo):
    """
    Transforma o balancete Excel do cliente Cooperlate.

    O arquivo é lido sem cabeçalho para permitir a identificação
    dinâmica da linha que contém as colunas Conta e Cta red.

    Layout de origem:
        A = Classificação e descrição
        B = Cta red
        C = Saldo Anterior
        D = Natureza do Saldo Anterior
        E = Débitos Mês
        F = Natureza dos Débitos
        G = Créditos Mês
        H = Natureza dos Créditos
        I = Saldo Mês
        J = Natureza do Saldo Mês
        K = Saldo Atual
        L = Natureza do Saldo Atual

    Layout de destino:
        A = Atividade
        B = Conta
        C = Nome
        D = Cód. Reduzido
        E = Saldo Anterior
        F = Débito
        G = Crédito
        H = Movimento
        I = Saldo Acumulado
    """
    nome_arquivo = os.path.basename(caminho_arquivo)
    extensao = os.path.splitext(caminho_arquivo)[1].lower()

    if extensao not in {".xls", ".xlsx"}:
        raise ValueError(
            f"O arquivo '{nome_arquivo}' não é um arquivo Excel válido."
        )

    engine = "xlrd" if extensao == ".xls" else "openpyxl"

    try:
        df_origem = pd.read_excel(
            caminho_arquivo,
            sheet_name=0,
            header=None,
            dtype=object,
            engine=engine
        )

    except Exception as erro:
        raise ValueError(
            f"Não foi possível ler o arquivo do cliente Cooperlate "
            f"'{nome_arquivo}'. Erro: {erro}"
        ) from erro

    if df_origem.shape[1] < 12:
        raise ValueError(
            f"O arquivo do cliente Cooperlate '{nome_arquivo}' possui "
            f"{df_origem.shape[1]} coluna(s), mas são necessárias "
            "pelo menos 12 colunas, de A até L."
        )

    indice_cabecalho = _localizar_cabecalho_cooperlate(
        df_origem,
        nome_arquivo
    )

    # Mantém somente as linhas posteriores ao cabeçalho.
    df_origem = df_origem.iloc[
        indice_cabecalho + 1:
    ].copy()

    # Remove linhas totalmente vazias, inclusive as linhas em branco
    # existentes entre as contas.
    df_origem.dropna(
        how="all",
        inplace=True
    )

    df_origem.reset_index(
        drop=True,
        inplace=True
    )

    registros_validos = []

    for _, linha in df_origem.iterrows():
        conta, descricao = _separar_conta_descricao_cooperlate(
            linha.iloc[0]
        )

        # Linhas sem classificação e descrição válidas são ignoradas.
        # Isso também evita processar cabeçalhos repetidos, totais sem
        # conta e outras informações complementares.
        if not conta or not descricao:
            continue

        codigo_reduzido = _normalizar_texto_cooperlate(
            linha.iloc[1]
        )

        registros_validos.append({
            "Conta": conta,
            "Nome": descricao,
            "Cód. Reduzido": codigo_reduzido,
            "Saldo Anterior": _aplicar_natureza_cooperlate(
                linha.iloc[2],
                linha.iloc[3]
            ),
            "Débito": abs(
                _converter_numero_cooperlate(
                    linha.iloc[4]
                )
            ),
            "Crédito": abs(
                _converter_numero_cooperlate(
                    linha.iloc[6]
                )
            ),
            "Saldo Acumulado": _aplicar_natureza_cooperlate(
                linha.iloc[10],
                linha.iloc[11]
            ),
        })

    if not registros_validos:
        raise ValueError(
            f"Nenhuma conta válida foi encontrada no arquivo "
            f"do cliente Cooperlate '{nome_arquivo}'."
        )

    df_registros = pd.DataFrame(
        registros_validos
    )

    df_destino = pd.DataFrame(
        index=df_registros.index
    )

    # Coluna A: valor fixo.
    df_destino["Atividade"] = "Geral"

    # Coluna B: classificação extraída da origem A.
    df_destino["Conta"] = df_registros["Conta"]

    # Coluna C: descrição extraída da origem A.
    df_destino["Nome"] = df_registros["Nome"]

    # Coluna D: origem B.
    df_destino["Cód. Reduzido"] = (
        df_registros["Cód. Reduzido"]
    )

    # Coluna E: origem C com natureza da origem D.
    df_destino["Saldo Anterior"] = (
        df_registros["Saldo Anterior"]
        .astype(float)
    )

    # Coluna F: origem E.
    df_destino["Débito"] = (
        df_registros["Débito"]
        .astype(float)
    )

    # Coluna G: origem G.
    df_destino["Crédito"] = (
        df_registros["Crédito"]
        .astype(float)
    )

    # Coluna H: Débito menos Crédito.
    df_destino["Movimento"] = (
        df_destino["Débito"]
        - df_destino["Crédito"]
    )

    # Coluna I: origem K com natureza da origem L.
    df_destino["Saldo Acumulado"] = (
        df_registros["Saldo Acumulado"]
        .astype(float)
    )

# ==============================================================================
# TRANSFORMAÇÃO DO BALANCETE AURIVERDE
# ==============================================================================

def _localizar_cabecalho_auriverde(df_origem, nome_arquivo):
    """
    Localiza a linha de cabeçalho do balancete.
    O cabeçalho deve apresentar 'Conta' na primeira coluna.
    """
    for indice in df_origem.index:
        valor_coluna_a = str(df_origem.iloc[indice, 0]).strip().upper()
        if valor_coluna_a == "CONTA":
            return indice

    raise ValueError(
        f"Não foi possível localizar o cabeçalho no arquivo '{nome_arquivo}'."
    )

def transformar_balancete_auriverde(caminho_arquivo):
    """
    Transforma o balancete Excel do cliente Auriverde.

    O cabeçalho é deslocado em relação aos dados. O Saldo Anterior real 
    está ausente, sendo calculado matematicamente a partir do Saldo Atual e movimentações.
    """
    nome_arquivo = os.path.basename(caminho_arquivo)
    extensao = os.path.splitext(caminho_arquivo)[1].lower()

    if extensao not in {".xls", ".xlsx"}:
        raise ValueError(
            f"O arquivo '{nome_arquivo}' não é um arquivo Excel válido."
        )

    engine = "xlrd" if extensao == ".xls" else "openpyxl"

    try:
        df_origem = pd.read_excel(
            caminho_arquivo,
            sheet_name=0,
            header=None,
            dtype=object,
            engine=engine
        )

    except Exception as erro:
        raise ValueError(
            f"Não foi possível ler o arquivo do cliente Auriverde "
            f"'{nome_arquivo}'. Erro: {erro}"
        ) from erro

    if df_origem.shape[1] < 7:
        raise ValueError(
            f"O arquivo do cliente Auriverde '{nome_arquivo}' possui "
            f"{df_origem.shape[1]} coluna(s), mas são necessárias "
            "pelo menos 7 colunas."
        )

    indice_cabecalho = _localizar_cabecalho_auriverde(
        df_origem,
        nome_arquivo
    )

    # Mantém apenas as linhas de dados (após o cabeçalho)
    df_origem = df_origem.iloc[
        indice_cabecalho + 1:
    ].copy()

    # Remove linhas completamente vazias
    df_origem.dropna(
        how="all",
        inplace=True
    )

    # Mantém apenas linhas que tenham a Conta (Coluna A) preenchida
    conta_preenchida = (
        df_origem.iloc[:, 0]
        .fillna("")
        .astype(str)
        .str.strip()
        .ne("")
    )

    df_origem = df_origem.loc[
        conta_preenchida
    ].copy()

    df_origem.reset_index(
        drop=True,
        inplace=True
    )

    if df_origem.empty:
        raise ValueError(
            f"O arquivo do cliente Auriverde '{nome_arquivo}' não possui "
            "linhas válidas para tabulação."
        )

    df_destino = pd.DataFrame(
        index=df_origem.index
    )

    # Coluna A: valor fixo
    df_destino["Atividade"] = "Geral"

    # Coluna B: Conta (Origem A)
    df_destino["Conta"] = (
        df_origem.iloc[:, 0]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    # Coluna D: Cód. Reduzido (Origem B com formatação para evitar .0)
    def parse_reduzido(x):
        if pd.isna(x) or str(x).strip() == "":
            return ""
        if isinstance(x, float) and x.is_integer():
            return str(int(x))
        return str(x).strip()

    df_destino["Cód. Reduzido"] = df_origem.iloc[:, 1].apply(parse_reduzido)

    # Coluna C: Nome (Origem C que caiu como Saldo Anterior nos títulos originais)
    df_destino["Nome"] = (
        df_origem.iloc[:, 2]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    # Valores numéricos puros (Débito e Crédito)
    debito = pd.to_numeric(
        df_origem.iloc[:, 3],
        errors="coerce"
    ).fillna(0.0)

    credito = pd.to_numeric(
        df_origem.iloc[:, 4],
        errors="coerce"
    ).fillna(0.0)

    # Saldo Atual
    saldo_atual = pd.to_numeric(
        df_origem.iloc[:, 6],
        errors="coerce"
    ).fillna(0.0)

    # Cálculos dinâmicos
    movimento = debito - credito
    saldo_anterior = saldo_atual - movimento

    df_destino["Saldo Anterior"] = saldo_anterior.round(2)
    df_destino["Débito"] = debito.round(2)
    df_destino["Crédito"] = credito.round(2)
    df_destino["Movimento"] = movimento.round(2)
    df_destino["Saldo Acumulado"] = saldo_atual.round(2)

# ==============================================================================
# TRANSFORMAÇÃO DO BALANCETE COOABRIEL (BIAGRE)
# ==============================================================================

def _aplicar_quebra_numero(base, complemento):
    """
    Junta a base do número com a parte que 'vazou' para a linha de baixo,
    injetando as casas decimais corretamente caso tenham sido omitidas.
    """
    base_str = str(base).strip() if pd.notna(base) else ""
    comp_str = str(complemento).strip() if pd.notna(complemento) else ""
    
    if not comp_str:
        return base_str
        
    if "," in base_str:
        # Se a base já tem vírgula, ex: '1.234,5' + '9' = '1.234,59'
        return base_str + comp_str
    elif "." in base_str:
        partes = base_str.split('.')
        # Se o ponto for o decimal (ex: '1684731931.1') junta direto
        if len(partes[-1]) <= 2:
            return base_str + comp_str
        else:
            # Se for separador de milhar (ex '1.887.918.607'), injeta a vírgula
            return base_str + ",0" + comp_str
    else:
        # Se não tem ponto nem vírgula, a base é inteira (ex: '1887918607') 
        # e o vazamento é o decimal que faltava.
        return base_str + ".0" + comp_str

def _converter_numero_cooabriel(valor):
    """
    Converte valores monetários do cliente Cooabriel para float.
    Trata formatações como '2.892.239.571,50' ou '0,00'.
    """
    if pd.isna(valor):
        return 0.0

    if isinstance(valor, (int, float)):
        return float(valor)

    texto = str(valor).strip()

    if not texto:
        return 0.0

    texto = texto.replace("\xa0", "").replace(" ", "")

    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")

    try:
        return float(texto)
    except (ValueError, TypeError):
        return 0.0

def _aplicar_natureza_cooabriel(valor, natureza):
    """
    Aplica o sinal do saldo conforme a natureza contábil.
    Regras: D = positivo, C = negativo.
    """
    numero = _converter_numero_cooabriel(valor)

    if pd.isna(natureza):
        return numero

    natureza_texto = str(natureza).strip().upper()

    if natureza_texto == "D":
        return abs(numero)
    if natureza_texto == "C":
        return -abs(numero)

    return numero

def transformar_balancete_cooabriel(caminho_arquivo):
    """
    Transforma o balancete Excel do cliente Cooabriel.

    Este layout possui um problema crônico de exportação onde 
    valores numéricos e nomes extensos de contas quebram para a linha de baixo.
    O código reconstrói os registros "órfãos" colando-os na linha principal.
    """
    nome_arquivo = os.path.basename(caminho_arquivo)
    extensao = os.path.splitext(caminho_arquivo)[1].lower()

    if extensao not in {".xls", ".xlsx"}:
        raise ValueError(f"O arquivo '{nome_arquivo}' não é um arquivo Excel válido.")

    engine = "xlrd" if extensao == ".xls" else "openpyxl"

    try:
        df_raw = pd.read_excel(
            caminho_arquivo, 
            sheet_name=0, 
            header=None, 
            dtype=str, 
            engine=engine
        )
    except Exception as erro:
        raise ValueError(f"Não foi possível ler o arquivo '{nome_arquivo}'. Erro: {erro}")

    if df_raw.shape[1] < 9:
        raise ValueError(
            f"O arquivo '{nome_arquivo}' possui {df_raw.shape[1]} coluna(s), "
            "mas o layout Cooabriel exige 9 colunas (Conta, Chave, Descrição, SA, Nat, Déb, Créd, SF, Nat)."
        )

    registros_corrigidos = []
    ultimo_registro = None

    # Algoritmo de varredura e reconstrução de quebras de linha
    for i, row in df_raw.iterrows():
        conta = str(row[0]).strip() if pd.notna(row[0]) else ""
        desc = str(row[2]).strip() if pd.notna(row[2]) else ""

        # Ignora cabeçalhos principais
        if conta in ["Conta", "COOP AGRARIA DOS CAFEICULTORES DE SAO GABRIEL", 
                     "Balancete de Verificação"] or conta.startswith("CNPJ:"):
            continue

        # Se a Conta está vazia, esta linha pode ser lixo de paginação ou a metade de uma linha cortada
        if not conta:
            col7 = str(row[7]).strip() if pd.notna(row[7]) else ""
            
            # Ignora quebras de página literais
            if "FOLHA:" in col7:
                continue

            # Reconstrução dos dados vazados na linha inferior
            if ultimo_registro is not None:
                if pd.notna(row[3]) and str(row[3]).strip():
                    ultimo_registro["Saldo Anterior"] = _aplicar_quebra_numero(ultimo_registro["Saldo Anterior"], row[3])
                if pd.notna(row[5]) and str(row[5]).strip():
                    ultimo_registro["Débitos"] = _aplicar_quebra_numero(ultimo_registro["Débitos"], row[5])
                if pd.notna(row[6]) and str(row[6]).strip():
                    ultimo_registro["Créditos"] = _aplicar_quebra_numero(ultimo_registro["Créditos"], row[6])
                if pd.notna(row[7]) and str(row[7]).strip():
                    ultimo_registro["Saldo Final"] = _aplicar_quebra_numero(ultimo_registro["Saldo Final"], row[7])
                
                # Se o nome quebrou (Ex: TITULOS E VA), junta com um espaço
                if desc:
                    ultimo_registro["Nome"] += " " + desc
            continue

        # Linha contábil principal identificada
        record = {
            "Conta": conta,
            "Chave": str(row[1]).strip() if pd.notna(row[1]) else "",
            "Nome": desc,
            "Saldo Anterior": str(row[3]).strip() if pd.notna(row[3]) else "0",
            "Nat SA": str(row[4]).strip() if pd.notna(row[4]) else "",
            "Débitos": str(row[5]).strip() if pd.notna(row[5]) else "0",
            "Créditos": str(row[6]).strip() if pd.notna(row[6]) else "0",
            "Saldo Final": str(row[7]).strip() if pd.notna(row[7]) else "0",
            "Nat SF": str(row[8]).strip() if pd.notna(row[8]) else ""
        }
        registros_corrigidos.append(record)
        ultimo_registro = record

    if not registros_corrigidos:
        raise ValueError(f"Nenhuma conta contábil foi encontrada no arquivo '{nome_arquivo}'.")

    # Transforma os dicionários limpos em DataFrame
    df_registros = pd.DataFrame(registros_corrigidos)
    df_destino = pd.DataFrame(index=df_registros.index)

    df_destino["Atividade"] = "Geral"
    df_destino["Conta"] = df_registros["Conta"]
    df_destino["Nome"] = df_registros["Nome"]
    df_destino["Cód. Reduzido"] = df_registros["Chave"]

    # Converte os números e aplica a natureza (Positivo = D, Negativo = C)
    saldo_anterior = pd.Series([
        _aplicar_natureza_cooabriel(val, nat) 
        for val, nat in zip(df_registros["Saldo Anterior"], df_registros["Nat SA"])
    ])
    
    debito = df_registros["Débitos"].apply(_converter_numero_cooabriel)
    credito = df_registros["Créditos"].apply(_converter_numero_cooabriel)
    
    saldo_acumulado = pd.Series([
        _aplicar_natureza_cooabriel(val, nat) 
        for val, nat in zip(df_registros["Saldo Final"], df_registros["Nat SF"])
    ])

    df_destino["Saldo Anterior"] = saldo_anterior
    df_destino["Débito"] = debito
    df_destino["Crédito"] = credito
    df_destino["Movimento"] = debito - credito
    df_destino["Saldo Acumulado"] = saldo_acumulado

# ==============================================================================
# TRANSFORMAÇÃO DO BALANCETE COOPERATIVA A1
# ==============================================================================

def _converter_numero_cooperativa_a1(valor):
    """
    Converte valores monetários do TXT da Cooperativa A1 para float.
    Trata formatações atípicas como '1948.655.803,10' ou negativos '-37.224.930,08'.
    """
    if not valor:
        return 0.0
    
    texto = str(valor).strip()
    if not texto:
        return 0.0

    # Remove o ponto de milhar e troca a vírgula decimal por ponto
    texto = texto.replace(".", "").replace(",", ".")
    
    try:
        return float(texto)
    except (ValueError, TypeError):
        return 0.0

def transformar_balancete_cooperativa_a1(caminho_arquivo):
    """
    Transforma o balancete TXT do cliente Cooperativa A1.
    
    Lê o arquivo texto linha a linha utilizando expressões regulares
    para driblar cabeçalhos e quebras de página. Calcula o Saldo Anterior
    dinamicamente (Saldo Atual - Movimento).
    """
    nome_arquivo = os.path.basename(caminho_arquivo)
    
    # Tenta múltiplas codificações para abrir o TXT legado
    codificacoes = ["utf-8-sig", "cp1252", "latin-1"]
    linhas = []
    
    for codificacao in codificacoes:
        try:
            with open(caminho_arquivo, mode="r", encoding=codificacao) as arquivo:
                linhas = arquivo.readlines()
            break
        except UnicodeDecodeError:
            continue
        except OSError as erro:
            raise ValueError(f"Não foi possível abrir o arquivo TXT '{nome_arquivo}'. Erro: {erro}")
            
    if not linhas:
        raise ValueError(f"Não foi possível ler o arquivo '{nome_arquivo}' ou ele está vazio.")

    # Regex para capturar: Conta, Nome, Debito, Credito, Movimento, Saldo Atual
    # Ex: ' 01 01 01   DISPONIVEL   593.578.129,84   642.748.012,78   -49.169.882,94   632.598.142,41'
    regex_linha_contabil = re.compile(
        r"^\s*([\d\s]+)\s+(.*?)\s+([-\d.,]+)\s+([-\d.,]+)\s+([-\d.,]+)\s+([-\d.,]+)\s*$"
    )
    
    registros = []
    
    for linha in linhas:
        # Ignora as quebras de página (form feed) comuns no layout
        linha_limpa = linha.replace("\x0c", "")
        
        match = regex_linha_contabil.match(linha_limpa)
        if match:
            conta_com_espacos = match.group(1).strip()
            nome_conta = match.group(2).strip()
            
            # Converte as colunas monetárias
            debito = _converter_numero_cooperativa_a1(match.group(3))
            credito = _converter_numero_cooperativa_a1(match.group(4))
            movimento = _converter_numero_cooperativa_a1(match.group(5))
            saldo_atual = _converter_numero_cooperativa_a1(match.group(6))
            
            # O Saldo Anterior é oculto no relatório, fazemos a engenharia reversa
            saldo_anterior = saldo_atual - movimento
            
            # O Código Reduzido será a conta sem os espaços
            codigo_reduzido = conta_com_espacos.replace(" ", "")
            
            registros.append({
                "Atividade": "Geral",
                "Conta": conta_com_espacos,
                "Nome": nome_conta,
                "Cód. Reduzido": codigo_reduzido,
                "Saldo Anterior": round(saldo_anterior, 2),
                "Débito": round(debito, 2),
                "Crédito": round(credito, 2),
                "Movimento": round(movimento, 2),
                "Saldo Acumulado": round(saldo_atual, 2)
            })

    if not registros:
        raise ValueError(f"Nenhuma conta contábil válida foi encontrada no arquivo '{nome_arquivo}'.")

    df_destino = pd.DataFrame(registros)

# ==============================================================================
# TRANSFORMAÇÃO DO BALANCETE LANGUIRU
# ==============================================================================

def transformar_balancete_languiru(caminho_arquivo):
    """
    Transforma o balancete Excel do cliente Languiru.
    
    Bypass: Usa uma leitura iterativa profunda via openpyxl para burlar 
    a trava de dimensões e extrair as células reais, garantindo a ordem 
    correta das colunas no padrão HUB.
    """
    import os
    import pandas as pd
    import warnings
    import re
    import openpyxl
    
    nome_arquivo = os.path.basename(caminho_arquivo)
    extensao = os.path.splitext(caminho_arquivo)[1].lower()

    if extensao not in {".xls", ".xlsx"}:
        raise ValueError(f"O arquivo '{nome_arquivo}' não é um arquivo Excel válido.")

    df_raw = None
    data_start_idx = None

    if extensao == ".xlsx":
        # Usa openpyxl nativo iterando célula a célula para ignorar as dimensões corrompidas do ERP
        try:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    data = []
                    # Lê de fato o que está preenchido, ignorando tags XML mentirosas
                    for row in ws.iter_rows(values_only=True):
                        data.append(row)
                    df_temp = pd.DataFrame(data)
                    
                    if df_temp.empty or df_temp.shape[1] < 7:
                        continue
                        
                    for i, row in df_temp.iterrows():
                        col0 = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
                        if re.match(r"^\.+(?:\s*)?\d", col0):
                            df_raw = df_temp
                            data_start_idx = i
                            break
                    if data_start_idx is not None:
                        break
        except Exception as e:
            raise ValueError(f"Não foi possível abrir o arquivo '{nome_arquivo}'. Erro: {e}")
            
    else:
        # Para arquivos .xls tradicionais usamos o xlrd
        engine = "xlrd"
        try:
            xls = pd.ExcelFile(caminho_arquivo, engine=engine)
            for sheet in xls.sheet_names:
                df_temp = pd.read_excel(xls, sheet_name=sheet, header=None, dtype=str)
                if df_temp.shape[1] >= 7:
                    for i, row in df_temp.iterrows():
                        col0 = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
                        if re.match(r"^\.+(?:\s*)?\d", col0):
                            df_raw = df_temp
                            data_start_idx = i
                            break
                if data_start_idx is not None:
                    break
        except Exception as e:
            raise ValueError(f"Não foi possível abrir o arquivo '{nome_arquivo}'. Erro: {e}")

    if df_raw is None or data_start_idx is None:
        raise ValueError(
            f"Não foi possível identificar os dados contábeis no arquivo '{nome_arquivo}'. "
            "O sistema varreu as abas e não encontrou o padrão de contas (ex: '..01') na primeira coluna."
        )

    # Isola apenas a área de dados a partir da linha encontrada (já exclui cabeçalhos)
    df_data = df_raw.iloc[data_start_idx:].copy()
    
    # Limpa o dataframe
    df_data.dropna(axis=1, how='all', inplace=True)
    df_data.dropna(subset=[df_data.columns[0]], inplace=True)
    df_data = df_data[df_data[df_data.columns[0]].astype(str).str.strip() != ""]
    df_data.reset_index(drop=True, inplace=True)

    if df_data.empty:
        raise ValueError(f"Nenhum dado contábil válido encontrado no arquivo '{nome_arquivo}'.")

    # Função lambda interna para cortar os pontos
    def _remover_dois_pontos_iniciais(val):
        texto = str(val).strip()
        if texto.startswith(".."):
            return texto[2:]
        elif texto.startswith("."): 
            return texto[1:]
        return texto

    conta_formatada = df_data.iloc[:, 0].apply(_remover_dois_pontos_iniciais)
    
    # Constrói o Dataframe final com a ORDEM DAS COLUNAS EXATA exigida pelo padrão
    df_destino = pd.DataFrame(index=df_data.index)
    df_destino["Atividade"] = "Geral"
    df_destino["Conta"] = conta_formatada
    df_destino["Nome"] = df_data.iloc[:, 2].fillna("").astype(str).str.strip()
    df_destino["Cód. Reduzido"] = df_data.iloc[:, 1].fillna("").astype(str).str.strip()
    
    # Higieniza a formatação contábil para cálculos precisos
    def _to_float(val):
        if pd.isna(val) or val is None: return 0.0
        texto = str(val).replace('R$', '').replace(' ', '').strip()
        if not texto: return 0.0
        if ',' in texto and '.' in texto:
            texto = texto.replace('.', '')
        texto = texto.replace(',', '.')
        try:
            return float(texto)
        except (ValueError, TypeError):
            return 0.0
            
    debito = df_data.iloc[:, 4].apply(_to_float)
    credito = df_data.iloc[:, 5].apply(_to_float)

    df_destino["Saldo Anterior"] = df_data.iloc[:, 3].apply(_to_float)
    df_destino["Débito"] = debito
    df_destino["Crédito"] = credito
    df_destino["Movimento"] = debito - credito
    df_destino["Saldo Acumulado"] = df_data.iloc[:, 6].apply(_to_float)
    
    return df_destino