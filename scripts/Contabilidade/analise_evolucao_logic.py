import os
import shutil
import tempfile
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter

try:
    from openpyxl.drawing.image import Image
    HAS_PILLOW = True
except ImportError:
    HAS_PILLOW = False

def parse_perc(val_str):
    if not val_str: return float('inf')
    v = val_str.replace('%', '').replace(',', '.').strip()
    try:
        return float(v) / 100.0
    except:
        return 0.0

def formatar_borda(celula):
    borda = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    celula.border = borda

def executar_analise_evolucao(caminho_entrada, opcoes, regras=None):
    if regras is None: regras = {}
    log_erros = []
    inclusao_inteligente = opcoes.get('inclusao_inteligente', True)
    is_reprocess = opcoes.get('is_reprocess', False)
    caminho_saida_desejado = opcoes.get('caminho_saida')
    
    # =========================================================================
    # 1. LEITURA E PADRONIZAÇÃO POR POSIÇÃO DE COLUNA (0 a 7 -> A a H)
    # =========================================================================
    if is_reprocess:
        caminho_pta_atual = opcoes.get('caminho_pta_reprocess')
        try:
            with pd.ExcelFile(caminho_pta_atual) as xl_params:
                if "Parametros" not in xl_params.sheet_names: raise ValueError(f"A aba 'Parametros' não foi encontrada.")
                df_plano_raw = xl_params.parse("Parametros", dtype=str, header=None)
        except Exception as e:
            raise ValueError(f"Erro ao ler os parâmetros do PTA editado. Detalhe: {e}")
    else:
        with pd.ExcelFile(caminho_entrada) as xl_params:
            nomes_possiveis = ['planodecontas', 'planodeconta', 'planoconta', 'parametros', 'cadastroparametros']
            abas_encontradas = [sht for sht in xl_params.sheet_names if str(sht).lower().replace(" ", "").replace("_", "").split("(")[0].strip() in nomes_possiveis]
            if not abas_encontradas: abas_encontradas = [xl_params.sheet_names[0]]
            df_plano_raw = xl_params.parse(abas_encontradas[-1], dtype=str, header=None)

    primeira_celula = str(df_plano_raw.iloc[0, 0]).upper().strip() if not df_plano_raw.empty else ""
    if 'CHAVE' in primeira_celula or 'CONTA' in primeira_celula:
        df_plano_raw = df_plano_raw.iloc[1:].reset_index(drop=True)

    # Garante 8 colunas (A a H) na leitura
    df_plano = pd.DataFrame()
    for col_idx in range(8):
        if col_idx < len(df_plano_raw.columns):
            df_plano[col_idx] = df_plano_raw.iloc[:, col_idx].astype(str)
        else:
            df_plano[col_idx] = ""

    df_plano = df_plano.fillna("")
    for c in range(8):
        df_plano[c] = df_plano[c].astype(str).str.strip()
        df_plano.loc[df_plano[c].str.lower().isin(['nan', 'null', 'none']), c] = ""

    df_plano[0] = df_plano[0].str.replace(r'[\.\-\s]', '', regex=True)
    df_plano = df_plano[df_plano[0] != ''].reset_index(drop=True)
    df_plano['Chave_Clean'] = df_plano[0]

    log_erros.append("Informação: Estrutura lida por posições (A-H). Pontuações (.) removidas internamente.")

    # =========================================================================
    # 2. LEITURA DOS BALANCETES
    # =========================================================================
    df_lista = []
    saldo_dezembro_dit = {}
    with pd.ExcelFile(caminho_entrada) as xl_dados:
        abas_meses = [sheet for sheet in xl_dados.sheet_names if str(sheet).isdigit() and 1 <= int(sheet) <= 12]
        if not abas_meses: raise ValueError("Nenhum balancete válido encontrado no arquivo original (abas 01 a 12).")
        for mes in abas_meses:
            df_mes_raw = xl_dados.parse(mes)
            if df_mes_raw.empty and len(df_mes_raw.columns) < 2: continue
            if len(df_mes_raw.columns) >= 9:
                df_mes = df_mes_raw.iloc[:, :9].copy()
                df_mes.columns = ['Atividade', 'Conta', 'Descrição', 'Cod. Reduzido', 'Saldo Anterior', 'Débito', 'Crédito', 'Movimento', 'Saldo Acumulado']
                for col_num in ['Saldo Anterior', 'Débito', 'Crédito', 'Movimento', 'Saldo Acumulado']:
                    df_mes[col_num] = pd.to_numeric(df_mes[col_num], errors='coerce').fillna(0.0)
                df_mes['Mês'] = int(mes)
                df_lista.append(df_mes)
                if mes == min(abas_meses, key=int):
                    df_mes['Conta_Format'] = df_mes['Conta'].astype(str).str.replace(r'[\.\-\s]', '', regex=True)
                    saldo_dezembro_dit = dict(zip(df_mes['Conta_Format'], df_mes['Saldo Anterior']))

    df_consolidado = pd.concat(df_lista, ignore_index=True)
    df_consolidado['Conta_Clean'] = df_consolidado['Conta'].astype(str).str.replace(r'[\.\-\s]', '', regex=True)
    df_consolidado['Conta'] = df_consolidado['Conta_Clean']
    meses_disponiveis = sorted(df_consolidado['Mês'].unique())
    ultimo_mes = meses_disponiveis[-1]
    num_meses_total = len(meses_disponiveis)

    # =========================================================================
    # 3. VERIFICAÇÃO DE CONTAS ÓRFÃS
    # =========================================================================
    contas_balancete = df_consolidado[['Cod. Reduzido', 'Conta_Clean', 'Descrição']].drop_duplicates(subset=['Conta_Clean'])
    contas_balancete = contas_balancete[contas_balancete['Conta_Clean'] != '']
    orfao_mask = ~contas_balancete['Conta_Clean'].isin(df_plano['Chave_Clean'])
    contas_orfas = contas_balancete[orfao_mask].copy()
    
    if not contas_orfas.empty and inclusao_inteligente:
        log_erros.append(f"Informação: {len(contas_orfas)} conta(s) órfãs identificadas e adicionadas ao plano de contas.")
        novas_linhas = pd.DataFrame({
            0: contas_orfas['Conta_Clean'],
            1: '',
            2: contas_orfas['Cod. Reduzido'],
            3: contas_orfas['Descrição'],
            4: '',
            5: '',
            6: '0',
            7: 'Adicionada por classificação automática',
            'Chave_Clean': contas_orfas['Conta_Clean']
        })
        df_plano = pd.concat([df_plano, novas_linhas], ignore_index=True)

    df_plano = df_plano.sort_values('Chave_Clean').reset_index(drop=True)
    
    # =========================================================================
    # 4. MOTOR DE RASTRO DE AUDITORIA (OTIMIZADO - PASSE ÚNICO)
    # =========================================================================
    if inclusao_inteligente or is_reprocess:
        chaves = df_plano['Chave_Clean'].tolist()
        
        # Predições vetoriais do Motor
        proxima_chave = chaves[1:] + ['']
        condicao_sintetica = [(str(prox).startswith(str(atual))) and (str(atual) != '') for atual, prox in zip(chaves, proxima_chave)]
        sint_an_motor = np.where(condicao_sintetica, 'S', 'A')
        prefixo = [str(c)[0] if c else '' for c in chaves]
        apr_motor = np.where(np.array(prefixo) == '1', 'A', np.where(np.array(prefixo) == '2', 'P', 'R'))
        
        # Loop unificado de alta performance (Avalia, Traduz e Grava na mesma passada)
        for idx in df_plano.index:
            u_sa = str(df_plano.at[idx, 4]).strip().upper()
            m_sa = sint_an_motor[idx]
            
            u_apr = str(df_plano.at[idx, 5]).strip().upper()
            m_apr = apr_motor[idx]
            
            obs_atual = str(df_plano.at[idx, 7]).strip()
            if obs_atual.lower() in ['nan', 'null', 'none']: obs_atual = ''
            
            mensagens = []
            
            # Análise e Tradução S/A
            if u_sa == '' or u_sa.lower() in ['nan', 'null', 'none']:
                df_plano.at[idx, 4] = m_sa
                mensagens.append(f"Auto-Classificada ({'Analítica' if m_sa == 'A' else 'Sintética'})")
            elif u_sa != m_sa:
                mensagens.append(f"Motor sugeriu S/A: {'Analítica' if m_sa == 'A' else 'Sintética'}")
                    
            # Análise e Tradução A/P/R
            if u_apr == '' or u_apr.lower() in ['nan', 'null', 'none']:
                df_plano.at[idx, 5] = m_apr
                if m_apr == 'A': mensagens.append("Auto-Classificada (Ativo)")
                elif m_apr == 'P': mensagens.append("Auto-Classificada (Passivo)")
                elif m_apr == 'R': mensagens.append("Auto-Classificada (Resultado)")
            elif u_apr != m_apr:
                if m_apr == 'A': mensagens.append("Motor sugeriu A/P/R: Ativo")
                elif m_apr == 'P': mensagens.append("Motor sugeriu A/P/R: Passivo")
                elif m_apr == 'R': mensagens.append("Motor sugeriu A/P/R: Resultado")
            
            # Concatenação e Gravação Blindada
            if mensagens:
                nova_msg = " | ".join(mensagens)
                if nova_msg not in obs_atual:
                    obs_atual = f"{obs_atual} | {nova_msg}" if obs_atual else nova_msg
                
                df_plano.at[idx, 7] = obs_atual

    df_plano[6] = [str(n) for n in range(1, len(df_plano) + 1)]
    
    # =========================================================================
    # 5. DIAGNÓSTICO MATEMÁTICO
    # =========================================================================
    df_pivot_mov = df_consolidado.pivot_table(index='Conta_Clean', columns='Mês', values='Movimento', aggfunc='sum').fillna(0)
    df_pivot_sld = df_consolidado.pivot_table(index='Conta_Clean', columns='Mês', values='Saldo Acumulado', aggfunc='sum').fillna(0)

    df_plano_a = df_plano[df_plano[4] == 'A']
    chaves_a = sorted(df_plano_a['Chave_Clean'].astype(str).tolist())
    pais_conflitantes = []
    for i in range(len(chaves_a) - 1):
        c_atual = chaves_a[i]
        c_prox = chaves_a[i+1]
        if c_prox.startswith(c_atual) and c_atual != c_prox: pais_conflitantes.append(c_atual)
    pais_conflitantes = sorted(list(set(pais_conflitantes)))
    if pais_conflitantes:
        log_erros.append(f"Conflito de Parametrização S/A: Foram identificadas {len(pais_conflitantes)} conta(s) Analíticas que possuem contas filhas também Analíticas.")

    todas_chaves_geral = df_plano['Chave_Clean'].tolist()
    contas_s_vazias = []
    for _, row_s in df_plano[df_plano[4] == 'S'].iterrows():
        chave_s = row_s['Chave_Clean']
        filhas_s = [c for c in todas_chaves_geral if c.startswith(chave_s)]
        if len(filhas_s) <= 1: contas_s_vazias.append(row_s[0])
    if contas_s_vazias:
        log_erros.append(f"Furo de Estrutura S/A: Identificadas {len(contas_s_vazias)} conta(s) parametrizada(s) como Sintética (S) sem subcontas cadastradas.")

    df_consolidado['Diff_Intramensal'] = df_consolidado['Saldo Anterior'] + df_consolidado['Movimento'] - df_consolidado['Saldo Acumulado']
    erros_math = df_consolidado[~np.isclose(df_consolidado['Diff_Intramensal'], 0, atol=0.01)]
    if not erros_math.empty:
        log_erros.append(f"Inconsistência Matemática Interna: {len(erros_math)} registro(s) apresentam divergência na equação (Saldo Anterior + Movimento != Saldo Acumulado).")

    for i in range(len(meses_disponiveis) - 1):
        m_ant = meses_disponiveis[i]; m_atu = meses_disponiveis[i+1]
        df_ant = df_consolidado[df_consolidado['Mês'] == m_ant][['Conta', 'Descrição', 'Saldo Acumulado']]
        df_atu = df_consolidado[df_consolidado['Mês'] == m_atu][['Conta', 'Saldo Anterior']]
        df_merge = pd.merge(df_ant, df_atu, on='Conta', how='inner')
        df_merge['Diff_Intermensal'] = df_merge['Saldo Acumulado'] - df_merge['Saldo Anterior']
        quebras = df_merge[~np.isclose(df_merge['Diff_Intermensal'], 0, atol=0.01)]
        if not quebras.empty:
            diff_liquida = quebras['Diff_Intermensal'].sum()
            log_erros.append(f"Quebra de Continuidade Intermensal (Mês {m_ant} -> Mês {m_atu}): {len(quebras)} conta(s) divergem R$ {diff_liquida:,.2f}.")

    for m in meses_disponiveis:
        soma_balancete = df_consolidado[df_consolidado['Mês'] == m]['Saldo Acumulado'].sum()
        if not np.isclose(soma_balancete, 0, atol=0.01):
            log_erros.append(f"Desbalanceamento Global da Origem: O balancete bruto do mês {m} não totaliza zero. Diferença total: R$ {soma_balancete:,.2f}.")

    def calc_base_val(chaves_input, col_idx_tipo, apr_tipo):
        if not chaves_input: return 0.0
        total = 0.0
        chaves_list = [c.strip() for c in str(chaves_input).split(',')]
        col_limpa = df_plano[col_idx_tipo].astype(str).str.replace(r'[\.\-\s0]$', '', regex=True).str.strip()
        for ch in chaves_list:
            ch_clean = ch.replace('.0', '').strip()
            contas_s = df_plano[(col_limpa == ch_clean) & (df_plano[4] == 'S') & (df_plano[5] == apr_tipo)][0].tolist()
            if contas_s:
                for c in contas_s:
                    if c in df_pivot_sld.index: total += df_pivot_sld.loc[c, ultimo_mes]
            else:
                contas_a = df_plano[(col_limpa == ch_clean) & (df_plano[4] == 'A') & (df_plano[5] == apr_tipo)][0].tolist()
                for c in contas_a:
                    if c in df_pivot_sld.index: total += df_pivot_sld.loc[c, ultimo_mes]
        return abs(total)

    bases_ativo = [calc_base_val(r['chave'], 1, 'A') for r in regras.get('ativo', [])] 
    bases_passivo = [calc_base_val(r['chave'], 0, 'P') for r in regras.get('passivo', [])] 
    bases_resultado = [calc_base_val(r['chave'], 1, 'R') for r in regras.get('resultado', [])] 

    df_plano = df_plano.drop(columns=['Chave_Clean'])

    # =========================================================================
    # 6. EXPORTAÇÃO
    # =========================================================================
    caminho_saida = caminho_saida_desejado
    base_saida_arq, ext = os.path.splitext(caminho_saida)
    counter = 2
    while os.path.exists(caminho_saida):
        try:
            with open(caminho_saida, 'r+'): break 
        except PermissionError:
            caminho_saida = f"{base_saida_arq}_v{counter}{ext}"
            counter += 1

    wb = Workbook()
    wb.remove(wb.active)
    
    f_aptos = Font(name='Aptos', size=11)
    f_aptos_bold = Font(name='Aptos', size=11, bold=True)
    f_check = Font(name='Aptos', size=11, bold=True, color="FF0000")
    fill_incl = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") 
    fill_ama = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") 
    fill_cabecalho = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid") 

    ws_param = wb.create_sheet("Parametros")
    ws_param.sheet_view.showGridLines = False
    
    colunas_padrao_export = ['Chave Cliente', 'Chave D&M', 'Classificação', 'Descrição', 'Sint./An.', 'At/Pas/Res', 'Indice', 'Observação']
    for c_idx, col_name in enumerate(colunas_padrao_export, 1): 
        cel = ws_param.cell(row=1, column=c_idx, value=col_name)
        cel.font = f_aptos_bold
        
    for r_idx, (_, row) in enumerate(df_plano.iterrows()):
        obs_text = str(row[7]) 
        
        # Filtro refinado para pintar a célula
        foi_ad = any(t in obs_text for t in ["Adicionada", "Auto-Classificada", "sugeriu"])
        
        for c_idx in range(8): 
            cel = ws_param.cell(row=r_idx+2, column=c_idx+1, value=str(row[c_idx]))
            cel.font = f_aptos
            
            # Pinta a célula de amarelo SOMENTE na coluna H (índice 7) se houve modificação
            if foi_ad and c_idx == 7: 
                cel.fill = fill_incl

    ws_param.auto_filter.ref = ws_param.dimensions
    for col in ws_param.columns: ws_param.column_dimensions[col[0].column_letter].width = 15
    ws_param.column_dimensions['D'].width = 35; 
    ws_param.column_dimensions['H'].width = 65 

    ws_hist = wb.create_sheet("Balancete_Histórico")
    ws_hist.sheet_view.showGridLines = False 
    
    ws_hist['G2'] = "Ativo Acumulado"; ws_hist['G3'] = "Passivo Acumulado"
    ws_hist['G4'] = "Resultado Mensal"; ws_hist['G5'] = "Resultado Acumulado"; ws_hist['G6'] = "Check"
    
    for r in range(2, 7): ws_hist[f'G{r}'].font = f_aptos_bold
    ws_hist['G6'].font = f_check

    col_cursor = 8
    for idx, m in enumerate(meses_disponiveis):
        l_mov = get_column_letter(col_cursor); l_sld = get_column_letter(col_cursor + 1)
        ws_hist[f'{l_mov}1'] = "Mês " + str(m)
        ws_hist[f'{l_mov}2'] = f'=SUMIFS({l_sld}:{l_sld}, $BB:$BB, "A", $BC:$BC, "A")'
        ws_hist[f'{l_mov}3'] = f'=SUMIFS({l_sld}:{l_sld}, $BB:$BB, "A", $BC:$BC, "P")'
        ws_hist[f'{l_mov}4'] = f'=SUMIFS({l_mov}:{l_mov}, $BB:$BB, "A", $BC:$BC, "R")'
        if idx == 0: ws_hist[f'{l_mov}5'] = f'={l_mov}4'
        else: ws_hist[f'{l_mov}5'] = f'={get_column_letter(col_cursor - 2)}5+{l_mov}4'
        ws_hist[f'{l_mov}6'] = f'={l_mov}2+{l_mov}3+{l_mov}5'
        col_cursor += 2

    headers_hist = ['Seleção', 'Atividade', 'Chave', 'Conta', 'Descrição', 'Cod. Reduzido', 'Acum. Dez']
    for m in meses_disponiveis: headers_hist.extend(['Movimento', 'Saldo Acumulado'])
    for c_idx, h_text in enumerate(headers_hist, 1):
        cel = ws_hist.cell(row=8, column=c_idx, value=h_text)
        cel.font = f_aptos_bold

    linha_inicio = 9
    for r_idx, (_, row_p) in enumerate(df_plano.iterrows()):
        current_row = linha_inicio + r_idx
        conta_cod = str(row_p[0]).strip() 
        selecao_val = ""; cols_amarelas = []
        
        if row_p[4] == 'A': 
            saldo_final_abs = abs(df_pivot_sld.loc[conta_cod, ultimo_mes]) if conta_cod in df_pivot_sld.index else 0.0
            if row_p[5] == 'A' and opcoes.get('ativo', False): 
                for i, r_dict in enumerate(regras.get('ativo', [])):
                    if not r_dict['min']: continue
                    b_a = bases_ativo[i]
                    if (b_a * parse_perc(r_dict['min'])) < saldo_final_abs <= (b_a * parse_perc(r_dict['max'])):
                        selecao_val = f"X-A{i+1}"; cols_amarelas.append(1); break
            elif row_p[5] == 'P' and opcoes.get('passivo', False):
                for i, r_dict in enumerate(regras.get('passivo', [])):
                    if not r_dict['min']: continue
                    b_p = bases_passivo[i]
                    if (b_p * parse_perc(r_dict['min'])) < saldo_final_abs <= (b_p * parse_perc(r_dict['max'])):
                        selecao_val = f"X-P{i+1}"; cols_amarelas.append(1); break
            elif row_p[5] == 'R' and opcoes.get('resultado', False):
                movs = [df_pivot_mov.loc[conta_cod, m] if conta_cod in df_pivot_mov.index else 0.0 for m in meses_disponiveis]
                media = sum(movs) / num_meses_total 
                for i, r_dict in enumerate(regras.get('resultado', [])):
                    if not r_dict['min']: continue
                    b_r = bases_resultado[i]; p_desv = parse_perc(r_dict['perc'])
                    if (b_r * parse_perc(r_dict['min'])) < saldo_final_abs <= (b_r * parse_perc(r_dict['max'])):
                        if media != 0:
                            matched = False; col_cursor_temp = 8
                            for m_v in movs:
                                if abs((m_v / media) - 1) >= p_desv:
                                    matched = True; cols_amarelas.append(col_cursor_temp)
                                col_cursor_temp += 2
                            if matched:
                                selecao_val = f"X-R{i+1}"
                                if 1 not in cols_amarelas: cols_amarelas.append(1)
                                break
        
        c_sel = ws_hist.cell(row=current_row, column=1, value=selecao_val)
        c_sel.font = f_aptos
        ws_hist.cell(row=current_row, column=57, value=selecao_val)
        if 1 in cols_amarelas: c_sel.fill = fill_ama
        
        dados_col = ["Geral", conta_cod, row_p[2], row_p[3], row_p[1], saldo_dezembro_dit.get(conta_cod, 0.0)]
        for ci, val in enumerate(dados_col, 2):
            cel = ws_hist.cell(row=current_row, column=ci, value=val)
            cel.font = f_aptos
            if ci == 7: cel.number_format = '#,##0.00'
        
        col_cursor = 8
        for m in meses_disponiveis:
            mov_val = df_pivot_mov.loc[conta_cod, m] if conta_cod in df_pivot_mov.index else 0.0
            sld_val = df_pivot_sld.loc[conta_cod, m] if conta_cod in df_pivot_sld.index else 0.0
            
            c_mov = ws_hist.cell(row=current_row, column=col_cursor, value=mov_val)
            c_mov.font = f_aptos; c_mov.number_format = '#,##0.00'
            if col_cursor in cols_amarelas: c_mov.fill = fill_ama
            
            c_sld = ws_hist.cell(row=current_row, column=col_cursor+1, value=sld_val)
            c_sld.font = f_aptos; c_sld.number_format = '#,##0.00'
            col_cursor += 2
            
        ws_hist.cell(row=current_row, column=54, value=row_p[4]) 
        ws_hist.cell(row=current_row, column=55, value=row_p[5]) 

    max_col_utilizada = 7 + (len(meses_disponiveis) * 2)
    for r in range(1, 8):
        for c in range(1, max_col_utilizada + 1):
            ws_hist.cell(row=r, column=c).fill = fill_cabecalho

    meses_nomes_dict = {1:"Janeiro", 2:"Fevereiro", 3:"Março", 4:"Abril", 5:"Maio", 6:"Junho", 7:"Julho", 8:"Agosto", 9:"Setembro", 10:"Outubro", 11:"Novembro", 12:"Dezembro"}
    
    ws_hist.column_dimensions['G'].width = 21.5 
    
    col_cursor = 8
    for m in meses_disponiveis:
        l_mov = get_column_letter(col_cursor)
        l_sld = get_column_letter(col_cursor + 1)
        
        ws_hist.cell(row=1, column=col_cursor, value=meses_nomes_dict.get(m, f"Mês {m}")).font = f_aptos_bold
        ws_hist.column_dimensions[l_mov].width = 20
        
        if m != ultimo_mes: ws_hist.column_dimensions[l_sld].hidden = True
        
        for r in range(2, 6): ws_hist.cell(row=r, column=col_cursor).font = f_aptos; ws_hist.cell(row=r, column=col_cursor).number_format = '#,##0.00'
        ws_hist.cell(row=6, column=col_cursor).font = f_check; ws_hist.cell(row=6, column=col_cursor).number_format = '#,##0.00'
        col_cursor += 2

    max_col_let = get_column_letter(max_col_utilizada)
    ws_hist.auto_filter.ref = f"A8:{max_col_let}{ws_hist.max_row}"

    if HAS_PILLOW and os.path.exists("logo_aber.png"):
        try:
            img = Image("logo_aber.png")
            target_height = 130 
            if img.height > 0:
                ratio = target_height / img.height
                img.width = int(img.width * ratio)
                img.height = target_height
            ws_hist.add_image(img, 'C1')
        except: pass

    for col in ws_hist.columns:
        if col[0].column < 7: ws_hist.column_dimensions[get_column_letter(col[0].column)].width = 15
    ws_hist.column_dimensions['E'].width = 35 

    ws_log = wb.create_sheet("Log_Diagnostico")
    ws_log['A1'] = "Relatório de Diagnóstico Analítico"; ws_log['A1'].font = f_aptos_bold
    if log_erros:
        for i, erro in enumerate(log_erros, start=3): ws_log.cell(row=i, column=1, value=erro).font = f_aptos
    else:
        ws_log.cell(row=3, column=1, value="Nenhuma inconsistência primária identificada no processamento estrutural.").font = f_aptos

    wb.save(caminho_saida)
    wb.close()
    
    return caminho_saida, log_erros


# ======================================================================================
# GERAÇÃO DE PTA: EXCEL SEPARADO + ABA "TABELAS"
# ======================================================================================
def gerar_ptas_excel(caminho_origem, caminho_destino):
    temp_dir = tempfile.gettempdir()
    temp_path = os.path.join(temp_dir, "temp_leitura_pta.xlsx")
    
    try: shutil.copy2(caminho_origem, temp_path)
    except Exception as e: raise ValueError(f"Não foi possível ler o arquivo original. Erro: {e}")
        
    wb_orig = load_workbook(temp_path)
    if "Balancete_Histórico" not in wb_orig.sheetnames:
        wb_orig.close(); os.remove(temp_path)
        raise ValueError("A aba 'Balancete_Histórico' não foi encontrada no arquivo base.")
        
    ws_hist = wb_orig["Balancete_Histórico"]
    ptas = {'A': [], 'P': [], 'R': []}
    justificadas = []
    todas_selecionadas = [] 
    
    num_meses = 0
    for c in range(8, ws_hist.max_column + 1, 2):
        val_header = str(ws_hist.cell(row=8, column=c).value or "").strip()
        if "Movimento" in val_header: num_meses += 1
        else: break
            
    for r in range(9, ws_hist.max_row + 1):
        tag_visivel = str(ws_hist.cell(row=r, column=1).value or "").strip().upper()
        tag_oculta = str(ws_hist.cell(row=r, column=57).value or "").strip().upper()
        
        if not tag_oculta and tag_visivel.startswith('X-'): tag_oculta = tag_visivel
            
        saldos_12 = [0.0] * 12
        meses_pintados = []
        
        for m in range(num_meses):
            c_mov = 8 + (m * 2)
            v = ws_hist.cell(row=r, column=c_mov).value
            saldos_12[m] = v if v is not None else 0.0
            
            cel_fill = ws_hist.cell(row=r, column=c_mov).fill
            if cel_fill and cel_fill.start_color.index != "00000000": 
                meses_pintados.append(m)
                
        col_ult_sld = 8 + ((num_meses - 1) * 2) + 1 if num_meses > 0 else 9
        val_acum = ws_hist.cell(row=r, column=col_ult_sld).value
                
        dados_conta = {
            'chave': ws_hist.cell(row=r, column=3).value,
            'desc': ws_hist.cell(row=r, column=5).value,
            'regra': tag_oculta.replace('X-', ''),
            'saldos': saldos_12, 
            'acumulado': val_acum if val_acum is not None else 0.0, 
            'pintura': meses_pintados
        }
        
        if tag_oculta.startswith('X-'):
            if tag_visivel == tag_oculta:
                todas_selecionadas.append(dados_conta)
                if 'A' in tag_oculta: ptas['A'].append(dados_conta)
                elif 'P' in tag_oculta: ptas['P'].append(dados_conta)
                elif 'R' in tag_oculta: ptas['R'].append(dados_conta)
            else: justificadas.append(dados_conta)

    wb_orig.close()
    try: os.remove(temp_path)
    except: pass

    wb_pta = Workbook()
    
    f_pta_padrao = Font(name='Cambria', size=11)
    f_pta_bold = Font(name='Cambria', size=11, bold=True)
    fill_ama = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") 
    fill_cinza = PatternFill(start_color="CBCBCB", end_color="CBCBCB", fill_type="solid")

    def construir_aba_pta(tipo, dados):
        ws = wb_pta.create_sheet(f"PTA - {tipo}")
        ws.sheet_view.showGridLines = False
        ws['A1'] = "DICKEL & MAFFI - AUDITORIA E CONSULTORIA SS"; ws['A1'].font = f_pta_bold
        ws['A2'] = "Rua Dr. Mario Totta, 714 - 3º andar"; ws['A2'].font = f_pta_padrao
        ws['A3'] = "Porto Alegre - RS"; ws['A3'].font = f_pta_padrao
        
        if tipo in ['A', 'P']:
            ws['A5'] = "Cliente:"; ws['A5'].font = f_pta_padrao
            ws['A7'] = f"ANÁLISE EVOLUÇÃO CONTAS DE {'ATIVOS' if tipo == 'A' else 'PASSIVOS'}"; ws['A7'].font = f_pta_bold
            ws['D7'] = "Período:"; ws['D7'].font = f_pta_padrao
            ws['F7'] = "Etapa:"; ws['F7'].font = f_pta_padrao
            linha_header = 13
            
            meses = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
            cabs = ["REF", "Chave", "Conta"] + meses + ["Acumulado"]
            for c, text in enumerate(cabs, 1):
                ws.cell(row=linha_header, column=c, value=text).font = f_pta_bold; formatar_borda(ws.cell(row=linha_header, column=c))
                if c > 3: ws.cell(row=linha_header+1, column=c, value="R$").font = f_pta_bold
            
            r_cursor = linha_header + 2
            for i, d in enumerate(dados, 1):
                ws.cell(row=r_cursor, column=1, value=f"C{tipo}-{i}").font = f_pta_padrao
                ws.cell(row=r_cursor, column=2, value=d['chave']).font = f_pta_padrao
                ws.cell(row=r_cursor, column=3, value=d['desc']).font = f_pta_padrao
                for m_idx, v in enumerate(d['saldos']):
                    col_m = m_idx + 4; c_mov = ws.cell(row=r_cursor, column=col_m, value=v); c_mov.number_format = '#,##0.00'
                    c_mov.font = f_pta_padrao; formatar_borda(c_mov)
                ws.cell(row=r_cursor, column=16, value=d['acumulado']).number_format = '#,##0.00'
                ws.cell(row=r_cursor, column=16).font = f_pta_padrao; formatar_borda(ws.cell(row=r_cursor, column=16))
                r_cursor += 1
                
        else: 
            ws['A5'] = "Cliente:"; ws['A5'].font = f_pta_padrao
            ws['A6'] = "Período:"; ws['A6'].font = f_pta_padrao
            ws['A7'] = "Etapa:"; ws['A7'].font = f_pta_padrao
            
            ws['B9'] = "Plano de abordagem:"; ws['B9'].font = f_pta_padrao
            ws['C9'] = "Abordagens e conclusões:"; ws['C9'].font = f_pta_padrao
            ws['A10'] = "Auditor (es):"; ws['A10'].font = f_pta_padrao
            ws['A11'] = "Data:"; ws['A11'].font = f_pta_padrao
            ws['A12'] = "Revisor:"; ws['A12'].font = f_pta_padrao
            ws['A13'] = "Data:"; ws['A13'].font = f_pta_padrao
            
            for col_c in [1, 2, 3]: ws.cell(row=10, column=col_c).fill = fill_cinza; ws.cell(row=12, column=col_c).fill = fill_cinza
            ws['B11'].alignment = Alignment(horizontal='left'); ws['C11'].alignment = Alignment(horizontal='left')
            ws['B13'].alignment = Alignment(horizontal='left'); ws['C13'].alignment = Alignment(horizontal='left')

            ws['A15'] = "ANÁLISE EVOLUÇÃO CONTAS DE RESULTADO"; ws['A15'].font = f_pta_bold
            ws['A16'] = "Critério de Seleção das Contas: (NBC TA 530 CFC)"; ws['A16'].font = f_pta_padrao
            ws['A17'] = "As contas objeto das avaliações da auditoria foram selecionadas com base em critério que levou em consideração a representação percentual sobre o ingresso/receita acumulada, combinado com nível de variação em relação a sua média mensal. Essa seleção é identificada pelo código X-R1 a 8."
            ws['A17'].font = f_pta_padrao
            ws['A21'] = "Adicionalmente em relação as contas selecionadas conforme critério acima, podem ser excluídas contas mediante justificativa e incluídas contas mediante seleção manual através da observação visual do comportamento dos saldos ou por características que leva o auditor a julgar como adequada/necessário. Essa seleção é identificada pelo código X-RA ou X-RR."
            ws['A21'].font = f_pta_padrao
            ws.merge_cells("A17:P19"); ws.merge_cells("A21:P24")
            ws['A17'].alignment = Alignment(wrapText=True, vertical='top'); ws['A21'].alignment = Alignment(wrapText=True, vertical='top')
            
            linha_header = 26
            meses = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
            cabs = ["REF", "Chave", "Conta"] + meses + ["Acumulado", "Anexos"]
            for c, text in enumerate(cabs, 1):
                ws.cell(row=linha_header, column=c, value=text).font = f_pta_bold; formatar_borda(ws.cell(row=linha_header, column=c))
                if c > 3 and c < 17: ws.cell(row=linha_header+1, column=c, value="R$").font = f_pta_bold
                
            r_cursor = linha_header + 2
            for i, d in enumerate(dados, 1):
                ws.cell(row=r_cursor, column=1, value=f"CR-{i}").font = f_pta_padrao
                ws.cell(row=r_cursor, column=2, value=d['chave']).font = f_pta_padrao
                ws.cell(row=r_cursor, column=3, value=d['desc']).font = f_pta_padrao
                for m_idx, v in enumerate(d['saldos']):
                    col_m = m_idx + 4; c_mov = ws.cell(row=r_cursor, column=col_m, value=v); c_mov.number_format = '#,##0.00'
                    c_mov.font = f_pta_padrao; formatar_borda(c_mov)
                    if m_idx in d['pintura']: c_mov.fill = fill_ama
                ws.cell(row=r_cursor, column=16, value=d['acumulado']).number_format = '#,##0.00'
                ws.cell(row=r_cursor, column=16).font = f_pta_padrao; formatar_borda(ws.cell(row=r_cursor, column=16))
                formatar_borda(ws.cell(row=r_cursor, column=17))
                
                ws.cell(row=r_cursor+2, column=2, value="Objetivos de auditoria:").font = f_pta_padrao
                ws.cell(row=r_cursor+5, column=2, value="Plano de abordagem:").font = f_pta_padrao
                ws.cell(row=r_cursor+8, column=2, value="Critério de seleção dos registros:").font = f_pta_padrao
                ws.cell(row=r_cursor+11, column=2, value="Solicitações:").font = f_pta_padrao
                ws.cell(row=r_cursor+14, column=2, value="Abordagens e constatações:").font = f_pta_padrao
                ws.cell(row=r_cursor+17, column=2, value="Conclusões:").font = f_pta_padrao
                r_cursor += 22

        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 40
        for m in range(4, 17): ws.column_dimensions[get_column_letter(m)].width = 14

    if ptas['A']: construir_aba_pta('A', ptas['A'])
    if ptas['P']: construir_aba_pta('P', ptas['P'])
    if ptas['R']: construir_aba_pta('R', ptas['R'])

    if todas_selecionadas:
        ws_tab = wb_pta.create_sheet("Tabelas")
        ws_tab.sheet_view.showGridLines = False
        
        ws_tab['A1'] = "RESUMO GERAL DAS CONTAS SELECIONADAS"
        ws_tab['A1'].font = f_pta_bold
        
        meses = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
        headers_geral = ["Conta", "Nome da Conta"] + meses + ["Saldo Acumulado"]
        
        for c, text in enumerate(headers_geral, 1):
            cel = ws_tab.cell(row=3, column=c, value=text)
            cel.font = f_pta_bold
            formatar_borda(cel)
            
        r_cursor = 4
        for d in todas_selecionadas:
            ws_tab.cell(row=r_cursor, column=1, value=d['chave']).font = f_pta_padrao; formatar_borda(ws_tab.cell(row=r_cursor, column=1))
            ws_tab.cell(row=r_cursor, column=2, value=d['desc']).font = f_pta_padrao; formatar_borda(ws_tab.cell(row=r_cursor, column=2))
            
            for m_idx, v in enumerate(d['saldos']):
                c_mov = ws_tab.cell(row=r_cursor, column=m_idx+3, value=v)
                c_mov.number_format = '#,##0.00'
                c_mov.font = f_pta_padrao
                formatar_borda(c_mov)
                
            c_acum = ws_tab.cell(row=r_cursor, column=15, value=d['acumulado'])
            c_acum.number_format = '#,##0.00'
            c_acum.font = f_pta_padrao
            formatar_borda(c_acum)
            r_cursor += 1

        ws_tab.column_dimensions['A'].width = 15
        ws_tab.column_dimensions['B'].width = 40
        for m in range(3, 16): ws_tab.column_dimensions[get_column_letter(m)].width = 14

        r_cursor += 5
        
        for d in todas_selecionadas:
            r_header = r_cursor
            r_atual = r_cursor + 1
            r_anterior = r_cursor + 2
            
            ws_tab.cell(row=r_header, column=1, value="Conta").font = f_pta_bold; formatar_borda(ws_tab.cell(row=r_header, column=1))
            ws_tab.cell(row=r_header, column=2, value="Nome da Conta").font = f_pta_bold; formatar_borda(ws_tab.cell(row=r_header, column=2))
            for m_idx, m_name in enumerate(meses):
                cel = ws_tab.cell(row=r_header, column=m_idx+3, value=m_name)
                cel.font = f_pta_bold; formatar_borda(cel)
            ws_tab.cell(row=r_header, column=15, value="Saldo Acumulado").font = f_pta_bold; formatar_borda(ws_tab.cell(row=r_header, column=15))
                
            ws_tab.cell(row=r_atual, column=1, value=d['chave']).font = f_pta_padrao; formatar_borda(ws_tab.cell(row=r_atual, column=1))
            ws_tab.cell(row=r_atual, column=2, value=d['desc']).font = f_pta_padrao; formatar_borda(ws_tab.cell(row=r_atual, column=2))
            for m_idx, v in enumerate(d['saldos']):
                cel = ws_tab.cell(row=r_atual, column=m_idx+3, value=v)
                cel.number_format = '#,##0.00'; cel.font = f_pta_padrao; formatar_borda(cel)
            c_acum_atual = ws_tab.cell(row=r_atual, column=15, value=d['acumulado'])
            c_acum_atual.number_format = '#,##0.00'; c_acum_atual.font = f_pta_padrao; formatar_borda(c_acum_atual)
                
            ws_tab.cell(row=r_anterior, column=1, value="Ano Anterior").font = f_pta_bold; formatar_borda(ws_tab.cell(row=r_anterior, column=1))
            ws_tab.cell(row=r_anterior, column=2, value="").font = f_pta_padrao; formatar_borda(ws_tab.cell(row=r_anterior, column=2))
            for m_idx in range(12):
                cel = ws_tab.cell(row=r_anterior, column=m_idx+3, value="")
                cel.number_format = '#,##0.00'; cel.font = f_pta_padrao; formatar_borda(cel)
            c_acum_ant = ws_tab.cell(row=r_anterior, column=15, value="")
            c_acum_ant.number_format = '#,##0.00'; c_acum_ant.font = f_pta_padrao; formatar_borda(c_acum_ant)
            
            r_cursor += 4 

    if justificadas:
        ws_just = wb_pta.create_sheet("Justificadas")
        ws_just.sheet_view.showGridLines = False
        ws_just['A1'] = "DICKEL & MAFFI - CONTAS EXCLUÍDAS DA AMOSTRAGEM"; ws_just['A1'].font = f_pta_bold
        r_just = 3
        for d in justificadas:
            ws_just.cell(row=r_just, column=1, value=d['chave']).font = f_pta_bold
            ws_just.cell(row=r_just, column=2, value=d['desc']).font = f_pta_bold
            ws_just.cell(row=r_just+1, column=1, value=f"Regra Estourada Originalmente: X-{d['regra']}").font = f_pta_padrao
            ws_just.cell(row=r_just+2, column=1, value="Justificativa do Auditor:").font = f_pta_padrao
            ws_just.cell(row=r_just+3, column=1, value="____________________________________________________________________________________________________").font = f_pta_padrao
            r_just += 5
        ws_just.column_dimensions['A'].width = 25; ws_just.column_dimensions['B'].width = 50

    if "Sheet" in wb_pta.sheetnames:
        if len(wb_pta.sheetnames) > 1:
            del wb_pta["Sheet"]
        else:
            ws_aviso = wb_pta["Sheet"]
            ws_aviso.title = "Aviso"
            ws_aviso['A1'] = "Nenhuma conta foi selecionada (Coluna A vazia) para geração de PTAs."

    base_saida_arq, ext = os.path.splitext(caminho_destino)
    counter = 2
    while os.path.exists(caminho_destino):
        try:
            with open(caminho_destino, 'r+'): break 
        except PermissionError:
            caminho_destino = f"{base_saida_arq}_v{counter}{ext}"
            counter += 1

    wb_pta.save(caminho_destino)
    wb_pta.close()
    
    return caminho_destino